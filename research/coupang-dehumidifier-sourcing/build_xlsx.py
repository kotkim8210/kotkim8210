#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
쿠팡 제습기 아이템위너 소싱 시트 생성기.
- 다나와 매입가(실데이터)는 채워져 있고, 쿠팡 위너가/판매자수 등은 [입력] 셀로 비워둠.
- 마진 관련 열을 맨 앞에 배치. 위너가만 입력하면 마진/등급이 자동 계산되는 라이브 수식.
- 다나와 매입가만으로 계산되는 '손익분기 위너가 / 8%·12% 목표 위너가'를 채워
  쿠팡에서 실제 위너가를 보는 순간 즉시 판단 가능.
data.json 형식: {"generated_kst": "...", "models": [ {..}, ... ]}
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(HERE, "data.json"), encoding="utf-8"))
OUT = os.path.join(HERE, "쿠팡_제습기_소싱_마진시트.xlsx")

models = DATA["models"]
gen = DATA.get("generated_kst", "")

# ---- 비용 모델 (사장님 지침 §8) ----
FEE = 0.12 + 0.01 + 0.03 + 0.01  # 수수료12 + 결제1 + 위험3 + 운영1 = 0.17
# 목표순이익률 판매가: P*(1-FEE-r) = 매입 + 택배  => P = (매입+택배)/(1-FEE-r)
# 손익분기 r=0, 8%, 12%

# ---- 스타일 ----
NAVY = "1F3864"; BLUE = "2E5496"; YEL = "FFF2CC"; GREEN = "E2EFDA"; GREY = "F2F2F2"; ORANGE="FCE4D6"
hdr_font = Font(bold=True, color="FFFFFF", size=10, name="맑은 고딕")
cell_font = Font(size=10, name="맑은 고딕")
bold = Font(bold=True, size=10, name="맑은 고딕")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

wb = Workbook()

# =====================================================================
# Sheet 1: 소싱 마진 시트
# =====================================================================
ws = wb.active
ws.title = "소싱_마진시트"

# 열 정의: (헤더, 폭, 그룹색, 입력여부)
# 그룹: M=마진(앞), IN=쿠팡입력, DN=다나와실데이터, ID=식별, REF=참고
COLS = [
    ("순위", 6, "M", False),
    ("등급\n(자동)", 8, "M", False),
    ("예상마진(원)\n=위너가 입력시", 13, "M", False),
    ("마진율(%)", 10, "M", False),
    ("손익분기\n위너가(원)", 12, "M", False),
    ("8%마진\n목표위너가", 12, "M", False),
    ("12%마진\n목표위너가", 12, "M", False),
    ("① [입력]\n쿠팡 위너가격", 13, "IN", True),
    ("② 다나와\n배송포함 매입가", 14, "DN", False),
    ("③ [입력]\n판매자부담 택배비", 12, "IN", True),
    ("브랜드", 10, "ID", False),
    ("정확한 모델번호", 18, "ID", False),
    ("상품명", 30, "ID", False),
    ("용량(L)", 7, "ID", False),
    ("주요기능/사용처", 20, "ID", False),
    ("다나와\n판매처수", 8, "DN", False),
    ("다나와 검색/상품 URL", 34, "DN", False),
    ("④ [입력]\n신품 판매자수", 10, "IN", True),
    ("⑤ [입력]\n위너 배송형태", 12, "IN", True),
    ("⑥ [입력]\n쿠팡 상품평수", 10, "IN", True),
    ("⑦ [입력]\n최근리뷰(Y/N)", 10, "IN", True),
    ("⑧ [입력]\n2위 판매자가", 11, "IN", True),
    ("⑨ [입력]\n쿠팡 상품 URL", 26, "IN", True),
    ("쿠팡 인기신호(리드,미검증)", 30, "REF", False),
    ("위험/비고", 24, "REF", False),
]
GROUP_FILL = {"M": BLUE, "IN": NAVY, "DN": "375623", "ID": "595959", "REF": "7F6000"}

# 안내 배너
ws.merge_cells("A1:Y1")
c = ws["A1"]
c.value = ("⚠ 다나와 매입가·판매처수·모델번호 = 실제 조회 데이터.  "
           "쿠팡 위너가·판매자수·배송형태·상품평(①④⑤⑥⑦⑧⑨) = 쿠팡 봇차단으로 자동취득 불가 → "
           "로그인 상태에서 직접 확인해 노란칸에 입력하면 마진·등급 자동계산.  '사용법_및_한계' 탭 필독.")
c.font = Font(bold=True, color="9C0006", size=10, name="맑은 고딕")
c.fill = PatternFill("solid", fgColor="FFF2CC")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:Y2")
c2 = ws["A2"]
c2.value = f"조사기준(KST): {gen}  |  비용모델: 수수료12%+결제1%+위험3%+운영1%=판매가의 17% + 택배비  |  손익분기위너가=(매입+택배)/0.83"
c2.font = Font(italic=True, size=9, color="404040", name="맑은 고딕")
c2.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

HDR_ROW = 3
for j, (name, w, grp, is_in) in enumerate(COLS, start=1):
    cell = ws.cell(row=HDR_ROW, column=j, value=name)
    cell.font = hdr_font
    cell.fill = PatternFill("solid", fgColor=GROUP_FILL[grp])
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[HDR_ROW].height = 34
ws.freeze_panes = "H4"  # 마진열 + 식별 일부 고정 느낌 (위너가 입력열까지 보이게)

# 컬럼 인덱스 헬퍼
# 1순위,2등급,3마진,4마진율,5손익BEP,6 8%,7 12%,8위너가IN,9매입가,10택배IN,
# 11브랜드,12모델,13상품명,14용량,15기능,16판매처,17URL,18판매자IN,19배송IN,20평수IN,21최근IN,22_2위IN,23쿠팡URL,24신호,25위험
def L(i): return get_column_letter(i)

r0 = HDR_ROW + 1
for idx, m in enumerate(models):
    r = r0 + idx
    buy = m.get("danawaLowestPrice") or 0
    ship_default = 0  # 대형 제습기 다수 무료배송 가정; 택배비는 사용자가 실제값 입력
    # 위너가 입력셀 참조
    win = f"H{r}"; buycol = f"I{r}"; shipcol = f"J{r}"
    # 값 채우기
    ws.cell(row=r, column=1, value=idx+1)
    # 등급 자동 (쿠팡 입력 채워지면 확정). 참조: 판매자수R18? 실제 열: 신품판매자=col18, 위너배송=col19, 상품평=col20, 마진율=col4
    grade_formula = (
        f'=IF(OR(H{r}="",I{r}=""),"위너가입력필요",'
        f'IF(OR(R{r}="",S{r}="",T{r}=""),"쿠팡검증필요",'
        f'IF(OR(S{r}="로켓그로스",R{r}<2,T{r}<30,D{r}<5),"제외",'
        f'IF(AND(R{r}>=2,S{r}="일반",T{r}>=100,D{r}>=12),"S",'
        f'IF(AND(R{r}>=2,S{r}="일반",T{r}>=50,D{r}>=8),"A",'
        f'IF(AND(T{r}>=30,R{r}>=2,D{r}>=5),"B","C"))))))'
    )
    ws.cell(row=r, column=2, value=grade_formula)
    # 예상마진 = 위너가 - 매입 - 위너가*FEE - 택배
    ws.cell(row=r, column=3, value=f'=IF(H{r}="","",ROUND(H{r}-I{r}-H{r}*{FEE}-J{r},0))')
    # 마진율
    ws.cell(row=r, column=4, value=f'=IF(OR(H{r}="",H{r}=0),"",ROUND(C{r}/H{r}*100,1))')
    # 손익분기/8%/12% 목표 위너가 — 매입가만으로 계산되므로 정적값으로 채워 어떤 뷰어에서도 즉시 표시
    # (택배비 0 기준. 택배비 입력시 예상마진 C열은 자동 반영, 이 목표값은 참고치)
    if buy:
        ws.cell(row=r, column=5, value=round(buy/(1-FEE)))
        ws.cell(row=r, column=6, value=round(buy/(1-FEE-0.08)))
        ws.cell(row=r, column=7, value=round(buy/(1-FEE-0.12)))
    # 위너가 입력 (빈칸)
    ws.cell(row=r, column=8, value=None)
    # 매입가 (실데이터)
    ws.cell(row=r, column=9, value=buy if buy else None)
    # 택배비 입력 (기본 0, 안내주석)
    ws.cell(row=r, column=10, value=ship_default)
    ws.cell(row=r, column=11, value=m.get("brand",""))
    ws.cell(row=r, column=12, value=m.get("modelNumber",""))
    ws.cell(row=r, column=13, value=m.get("productName",""))
    ws.cell(row=r, column=14, value=m.get("capacityL",""))
    ws.cell(row=r, column=15, value=m.get("keyFeatures",""))
    ws.cell(row=r, column=16, value=m.get("sellerCount",""))
    ws.cell(row=r, column=17, value=m.get("danawaSearchUrl",""))
    # 쿠팡 입력열 18~23 비움
    for cc in range(18, 24):
        ws.cell(row=r, column=cc, value=None)
    ws.cell(row=r, column=24, value=m.get("coupangPopularitySignal",""))
    ws.cell(row=r, column=25, value=m.get("riskNotes","") or m.get("notes",""))

    # 스타일 적용
    for j in range(1, len(COLS)+1):
        cell = ws.cell(row=r, column=j)
        cell.border = border
        cell.font = cell_font
        grp = COLS[j-1][2]; is_in = COLS[j-1][3]
        if is_in:
            cell.fill = PatternFill("solid", fgColor=YEL)
        elif grp == "M":
            cell.fill = PatternFill("solid", fgColor=GREEN)
        elif grp == "DN":
            cell.fill = PatternFill("solid", fgColor="F2F7EE")
        else:
            cell.fill = PatternFill("solid", fgColor="FFFFFF")
        if j in (3,5,6,7,8,9,22):
            cell.alignment = right; cell.number_format = '#,##0'
        elif j in (4,):
            cell.alignment = right; cell.number_format = '0.0'
        elif j in (13,15,17,23,24,25):
            cell.alignment = left
        else:
            cell.alignment = center
    ws.row_dimensions[r].height = 30

# 위너가 입력칸에 주석
tip = Comment("쿠팡(로그인)에서 이 모델 상세페이지의 현재 아이템위너 판매가를 확인해 입력하세요. 입력 즉시 마진/등급 자동계산.", "소싱봇")
ws.cell(row=r0, column=8).comment = tip

last_row = r0 + len(models) - 1
# 조건부 서식 — 마진(C): 양수=초록, 0~10원 미만/손실=빨강 (마진 한눈에)
rng_c = f"C{r0}:C{last_row}"
ws.conditional_formatting.add(rng_c, CellIsRule(operator="greaterThanOrEqual", formula=["10"],
    fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True, name="맑은 고딕")))
ws.conditional_formatting.add(rng_c, CellIsRule(operator="lessThan", formula=["10"],
    fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", name="맑은 고딕")))
# 마진율(D): >=12 진초록, 8~12 초록, 5~8 노랑, <5 빨강
rng_d = f"D{r0}:D{last_row}"
ws.conditional_formatting.add(rng_d, CellIsRule(operator="greaterThanOrEqual", formula=["12"], fill=PatternFill("solid", fgColor="63BE7B")))
ws.conditional_formatting.add(rng_d, CellIsRule(operator="between", formula=["8","11.999"], fill=PatternFill("solid", fgColor="C6EFCE")))
ws.conditional_formatting.add(rng_d, CellIsRule(operator="between", formula=["5","7.999"], fill=PatternFill("solid", fgColor="FFEB9C")))
ws.conditional_formatting.add(rng_d, CellIsRule(operator="lessThan", formula=["5"], fill=PatternFill("solid", fgColor="FFC7CE")))
# 등급(B): S/A 진초록, B 노랑, C 회색, 제외 빨강
rng_b = f"B{r0}:B{last_row}"
ws.conditional_formatting.add(rng_b, FormulaRule(formula=[f'OR(B{r0}="S",B{r0}="A")'], fill=PatternFill("solid", fgColor="63BE7B"), font=Font(bold=True, color="FFFFFF", name="맑은 고딕")))
ws.conditional_formatting.add(rng_b, FormulaRule(formula=[f'B{r0}="B"'], fill=PatternFill("solid", fgColor="FFEB9C")))
ws.conditional_formatting.add(rng_b, FormulaRule(formula=[f'B{r0}="제외"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", name="맑은 고딕")))

# 입력 드롭다운
dv_ship = DataValidation(type="list", formula1='"일반,로켓그로스"', allow_blank=True)
dv_ship.error = "일반 또는 로켓그로스 중 선택"; dv_ship.prompt = "위너 배송형태를 선택하세요"
ws.add_data_validation(dv_ship); dv_ship.add(f"S{r0}:S{last_row}")
dv_rev = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
ws.add_data_validation(dv_rev); dv_rev.add(f"U{r0}:U{last_row}")

# =====================================================================
# Sheet 2: 사용법 및 한계
# =====================================================================
ws2 = wb.create_sheet("사용법_및_한계")
ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 110
guide = [
    ("H1", "⚠ 이 시트로 무엇이 되고, 무엇이 안 되는가 (반드시 읽기)"),
    ("P", ""),
    ("H2", "1. 자동으로 채워진 것 (실제 조회 데이터)"),
    ("P", "• 다나와 모델번호 / 최저가(매입가 후보) / 등록 판매처 수 — search.danawa.com 실제 조회값."),
    ("P", "• 이 값들엔 '다나와 검색 URL' 근거가 함께 들어 있습니다."),
    ("P", ""),
    ("H2", "2. 비워둔 것 (쿠팡 봇차단으로 자동취득 불가 — 노란칸 ①④⑤⑥⑦⑧⑨)"),
    ("P", "• 쿠팡 위너가격 / 신품 판매자 수 / 위너 배송형태(일반·로켓그로스) / 상품평 수 / 최근리뷰 / 2위가 / 상품 URL."),
    ("P", "• 이유: 이 환경에서 쿠팡은 403 차단 + 실브라우저도 연결리셋(Akamai 봇차단, 데이터센터 IP). 로그인 자격증명도 없음."),
    ("P", "• 사장님 지침(제외조건3 '판매자수 확인불가→제외/보류', 조사원칙 '추정금지')상, 저는 이 값을 지어내지 않습니다."),
    ("P", ""),
    ("H2", "3. 사용법 — 5분이면 한 행 완성"),
    ("P", "① 쿠팡(로그인)에서 모델번호로 검색 → 동일 모델 상세페이지 진입."),
    ("P", "② '다른 판매자 보기'로 신품 판매자 수·위너 배송형태·위너가·2위가 확인 → 노란칸 입력."),
    ("P", "③ 상품평 수·최근리뷰 날짜 확인 → 입력. 입력 즉시 '등급/예상마진/마진율' 자동 계산."),
    ("P", "④ 혹은 반대로: 이미 채워진 '손익분기/8%/12% 목표 위너가'를 보고, 쿠팡 실제 위너가가 그보다 높으면 진입 검토."),
    ("P", ""),
    ("H2", "4. 마진 계산식 (사장님 지침 §8, 보수적)"),
    ("P", "예상마진 = 쿠팡위너가 − 다나와매입가 − 쿠팡위너가×17%(수수료12+결제1+위험3+운영1) − 판매자부담 택배비"),
    ("P", "손익분기 위너가 = (매입가+택배비) ÷ 0.83   /   8%목표 = ÷0.75   /   12%목표 = ÷0.71"),
    ("P", "※ 택배비 기본 0원(대형 무료배송 가정). 실제 택배비/매입배송비 있으면 ③열에 입력."),
    ("P", ""),
    ("H2", "5. 등급 자동판정 로직 (쿠팡 입력 채워지면 확정)"),
    ("P", "제외: 위너=로켓그로스 or 판매자<2 or 상품평<30 or 마진율<5%"),
    ("P", "S: 판매자≥2·위너=일반·상품평≥100·마진율≥12   |   A: …·상품평≥50·마진율≥8   |   B: 상품평≥30·판매자≥2·마진율≥5"),
    ("P", ""),
    ("H2", "6. 쿠팡 데이터를 제대로 뽑는 3가지 정식 방법(택1)"),
    ("P", "(a) 사장님 PC/휴대폰(가정용 IP)+쿠팡 로그인으로 직접 확인 — 가장 정확, 무료."),
    ("P", "(b) 셀러 분석 SaaS: 아이템스카우트 / 판다랭크 / 셀러라이프 — 아이템위너 경쟁·판매자수·리뷰추이 제공(유료)."),
    ("P", "(c) 주거용 프록시+로그인 크롤러를 별도(비차단) 환경에서 운영 — 이 클라우드 환경에선 불가."),
    ("P", ""),
    ("H2", "7. 냉정한 결론 (사장님 6대 질문)"),
    ("P", "쿠팡 실측 없이 '지금 주문 발생 가능성/위너 승산/5%하락 방어/재고위험'은 확정 불가. 다나와 매입원가는 확보됨."),
    ("P", "→ 다음 액션: 이 시트 상위 후보 몇 개를 쿠팡 로그인으로 5분 검증 → 노란칸 채우면 즉시 S/A/B 자동 선별."),
]
rr = 1
for kind, text in guide:
    cell = ws2.cell(row=rr, column=2, value=text)
    if kind == "H1":
        cell.font = Font(bold=True, size=13, color="9C0006", name="맑은 고딕")
    elif kind == "H2":
        cell.font = Font(bold=True, size=11, color="1F3864", name="맑은 고딕")
    else:
        cell.font = Font(size=10, name="맑은 고딕")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rr += 1

# =====================================================================
# Sheet 3: 원본 다나와 데이터 (raw)
# =====================================================================
ws3 = wb.create_sheet("원본_다나와데이터")
raw_cols = ["세그먼트","브랜드","상품명","모델번호","용량(L)","다나와최저가","판매처수","출시연도","주요기능","다나와URL","비고"]
for j, h in enumerate(raw_cols, 1):
    c = ws3.cell(row=1, column=j, value=h); c.font = hdr_font; c.fill = PatternFill("solid", fgColor=NAVY); c.alignment=center; c.border=border
for i, m in enumerate(DATA.get("raw", models), start=2):
    vals = [m.get("segment",""), m.get("brand",""), m.get("productName",""), m.get("modelNumber",""),
            m.get("capacityL",""), m.get("danawaLowestPrice",""), m.get("sellerCount",""),
            m.get("releaseYear",""), m.get("keyFeatures",""), m.get("danawaSearchUrl",""), m.get("notes","")]
    for j, v in enumerate(vals, 1):
        c = ws3.cell(row=i, column=j, value=v); c.font=cell_font; c.border=border
        c.alignment = left if j in (3,9,10) else center
        if j == 6: c.number_format = '#,##0'
widths3 = [16,10,32,18,7,12,8,8,22,34,20]
for j,w in enumerate(widths3,1): ws3.column_dimensions[get_column_letter(j)].width = w
ws3.freeze_panes = "A2"

wb.save(OUT)
print("SAVED:", OUT)
print("models:", len(models))
