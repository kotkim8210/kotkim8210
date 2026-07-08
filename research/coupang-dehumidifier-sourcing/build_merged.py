#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""통합 소싱 마진시트 — 내 84개 조사 + 업로드 2개 파일 병합. 쿠팡 수수료 11.88% 기준 실마진."""
import json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

HERE=os.path.dirname(os.path.abspath(__file__))
MINE=json.load(open(f"{HERE}/data.json",encoding="utf-8"))["models"]
UP=json.load(open(f"{HERE}/uploaded.json",encoding="utf-8"))
OUT=f"{HERE}/쿠팡_제습기_통합_소싱마진시트.xlsx"
FEE=0.1188

def norm(m):
    if not m: return ""
    m=re.sub(r"\(.*?\)","",str(m)); return re.sub(r"\s","",m).upper()
def num(x):
    try: return float(str(x).replace(",","").strip())
    except: return None
def keyword(brand,cap):
    p=[]; c=num(cap)
    if c: p.append(f"{int(round(c))}L 제습기")
    if brand and str(brand).strip() and str(brand)!="None": p.append(f"{brand} 제습기")
    return " / ".join(p) if p else "제습기"

master={}
def ensure(k,brand):
    if k not in master:
        master[k]={"brand":brand,"model":"","cap":"","danawaBuy":None,"danawaUrl":"","coupangUrl":"",
                   "coupangPrice":None,"domesticSeller":"","sourcingNote":"","danawaSellers":"",
                   "reviews":"","winnerShip":"","sellerCountNew":"","signal":"","risk":"","src":set()}
    return master[k]

for m in MINE:
    k=norm(m.get("modelNumber"))
    if not k: continue
    r=ensure(k,m.get("brand",""))
    r["model"]=m.get("modelNumber",""); r["cap"]=m.get("capacityL","") or r["cap"]
    sp=num(m.get("sourcingPrice")); dl=num(m.get("danawaLowestPrice"))
    if sp and dl and sp>dl*1.12:   # 소싱가 이상치(판매처표 미로드 등) → 검색 최저가로 폴백
        r["danawaBuy"]=dl
        r["domesticSeller"]=(m.get("sourcingSeller","") or "")+" (가격 재확인 필요)"
        r["sourcingNote"]="⚠소싱가 이상치→다나와최저가 사용"
    else:
        if sp or dl: r["danawaBuy"]=sp or dl
        r["domesticSeller"]=m.get("sourcingSeller",""); r["sourcingNote"]=m.get("sourcingNote","")
    r["danawaUrl"]=m.get("sourcingUrl","") or m.get("danawaSearchUrl","") or r["danawaUrl"]
    r["danawaSellers"]=m.get("sellerCount","")
    if m.get("coupangPopularitySignal"): r["signal"]=m["coupangPopularitySignal"]
    if m.get("riskNotes"): r["risk"]=m["riskNotes"]
    r["src"].add("내조사(다나와84)")

for m in UP["f1"]:
    k=norm(m.get("model"))
    if not k: continue
    r=ensure(k,m.get("brand",""))
    if not r["model"]: r["model"]=re.sub(r"\s*\(.*?\)","",str(m.get("model","")))
    r["cap"]=r["cap"] or m.get("cap","")
    if not r["danawaBuy"] and num(m.get("danawa")): r["danawaBuy"]=num(m.get("danawa"))
    if not r["danawaUrl"] and m.get("danawaUrl"): r["danawaUrl"]=m["danawaUrl"]
    if not r["coupangUrl"] and m.get("coupangUrl"): r["coupangUrl"]=m["coupangUrl"]
    if m.get("memo") and not r["risk"]: r["risk"]=str(m.get("memo",""))[:60]
    r["src"].add("업로드①마진계산")

for m in UP["watch"]:
    k=norm(m.get("model"))
    if not k: continue
    r=ensure(k,m.get("brand",""))
    if not r["model"]: r["model"]=str(m.get("model",""))
    r["cap"]=r["cap"] or m.get("cap","")
    if not r["danawaBuy"] and num(m.get("danawa")): r["danawaBuy"]=num(m.get("danawa"))
    if not r["danawaUrl"] and m.get("danawaUrl"): r["danawaUrl"]=m["danawaUrl"]
    if not r["coupangUrl"] and m.get("coupangUrl"): r["coupangUrl"]=m["coupangUrl"]
    r["src"].add("업로드워치리스트")

for m in UP["cand"]:
    mk=m.get("model",""); disp=re.sub(r"\s*\(.*?\)","",str(mk))
    if norm(mk) and "미노출" not in str(mk) and "미확인" not in str(mk) and "PB" not in str(mk):
        k=norm(mk)
    else:
        k="UP::"+str(m.get("brand",""))+"::"+str(m.get("name",""))[:20]; disp="(모델미노출)"
    r=ensure(k,m.get("brand",""))
    if not r["model"]: r["model"]=disp
    r["cap"]=r["cap"] or m.get("cap","")
    if num(m.get("coupangPrice")): r["coupangPrice"]=num(m.get("coupangPrice"))
    if m.get("reviews") not in (None,""): r["reviews"]=m.get("reviews")
    if m.get("winnerShip"): r["winnerShip"]=m.get("winnerShip")
    if m.get("coupangUrl"): r["coupangUrl"]=m.get("coupangUrl")
    if not r["danawaBuy"] and num(m.get("danawa")): r["danawaBuy"]=num(m.get("danawa"))
    if not r["danawaUrl"] and m.get("danawaUrl"): r["danawaUrl"]=m.get("danawaUrl")
    ws_=str(m.get("winnerShip","")); st=str(m.get("status","")); rs=str(m.get("reason",""))
    tag=""
    if "로켓" in ws_: tag="[로켓형위너→제외] "
    elif "품절" in ws_ or "판매중지" in ws_: tag="[품절→제외] "
    elif st=="제외": tag="[업로드판정:제외] "
    if tag or rs: r["risk"]=(tag+rs)[:70]
    r["src"].add("업로드쿠팡실측")

rows=list(master.values())
for r in rows:
    r["keyword"]=keyword(r["brand"],r["cap"])
    if not r["danawaUrl"] and r["model"]:
        r["danawaUrl"]=f"https://search.danawa.com/dsearch.php?query={r['model']}"
    if not r["coupangUrl"] and r["model"] and r["model"]!="(모델미노출)":
        r["coupangUrl"]=f"https://www.coupang.com/np/search?q={r['model']}"

def est(r):
    if r["coupangPrice"] and r["danawaBuy"]:
        return r["coupangPrice"]-r["danawaBuy"]-r["coupangPrice"]*FEE
    return None
def sk(r):
    exc=("로켓" in str(r["winnerShip"])) or ("품절" in str(r["risk"])) or ("제외" in str(r["risk"]))
    e=est(r); has=r["coupangPrice"] is not None
    return (0 if (has and not exc) else (1 if has else 2), -(e if e is not None else -9e9), r["danawaBuy"] or 9e9)
rows.sort(key=sk)

# ===== workbook =====
NAVY="1F3864"; hdrf=Font(bold=True,color="FFFFFF",size=10,name="맑은 고딕")
cf=Font(size=10,name="맑은 고딕"); thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
rgt=Alignment(horizontal="right",vertical="center")
link=Font(size=9,color="0563C1",underline="single",name="맑은 고딕")

wb=Workbook(); ws=wb.active; ws.title="통합_소싱마진"
COLS=[("브랜드",10,"ID"),("키워드",18,"ID"),("모델명",20,"ID"),
 ("다나와 매입가\n(국내최저,원)",13,"BUY"),("다나와 URL",22,"URL"),("쿠팡 URL",22,"URL"),
 ("쿠팡 판매가\n(실측/입력)",12,"CP"),("예상마진(원)\n수수료11.88%",13,"MG"),("마진율\n(%)",8,"MG"),
 ("마진≥10원?",10,"MG"),("손익분기\n쿠팡가",11,"MG"),("택배비\n(입력)",9,"IN"),
 ("국내최저\n소싱처",13,"SRC"),("소싱 비고",20,"SRC"),("용량\n(L)",6,"ID"),("다나와\n판매처수",8,"DN"),
 ("쿠팡\n리뷰수",8,"CP"),("위너\n배송형태",12,"CP"),("신품판매자수\n(입력)",10,"IN"),
 ("데이터출처",16,"REF"),("쿠팡 인기신호/비고",26,"REF"),("위험/판정",26,"REF")]
GF={"ID":"595959","BUY":"BF8F00","URL":"7F6000","CP":"1F3864","MG":"2E5496","IN":"C00000","SRC":"806000","DN":"375623","REF":"7F6000"}
FILL={"ID":"F2F2F2","BUY":"FFF2CC","URL":"FFFFFF","CP":"DDEBF7","MG":"E2EFDA","IN":"FCE4D6","SRC":"FFF2CC","DN":"F2F7EE","REF":"FFFFFF"}
NC=len(COLS)
# 위치 기반 컬럼 문자 (백슬래시 회피)
Cbuy=get_column_letter(4); Cdurl=get_column_letter(5); Ccurl=get_column_letter(6); Ccp=get_column_letter(7)
Cmg=get_column_letter(8); Cmr=get_column_letter(9); Cjd=get_column_letter(10); Cbep=get_column_letter(11)
Ctb=get_column_letter(12); Cws=get_column_letter(18)

ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NC)
b=ws.cell(1,1,"통합 소싱·마진 시트 — 내 다나와 84종 + 업로드 2개 파일 병합 | 쿠팡 판매수수료 11.88% 적용, 예상마진 ≥10원만 '✅마진OK'")
b.font=Font(bold=True,color="FFFFFF",size=11,name="맑은 고딕"); b.fill=PatternFill("solid",fgColor=NAVY); b.alignment=lft
ws.row_dimensions[1].height=24
ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=NC)
b2=ws.cell(2,1,"예상마진 = 쿠팡판매가 − 다나와매입가 − 쿠팡판매가×11.88% − 택배비 | 파란셀 쿠팡가=업로드 실측값, 빈칸은 쿠팡 로그인으로 확인 입력 | 손익분기쿠팡가=매입가/0.8812")
b2.font=Font(italic=True,size=9,color="404040",name="맑은 고딕"); b2.alignment=lft
ws.row_dimensions[2].height=15
for j,(nm,w,gp) in enumerate(COLS,1):
    c=ws.cell(3,j,nm); c.font=hdrf; c.fill=PatternFill("solid",fgColor=GF[gp]); c.alignment=ctr; c.border=bd
    ws.column_dimensions[get_column_letter(j)].width=w
ws.row_dimensions[3].height=32
ws.freeze_panes="H4"

r0=4
for i,r in enumerate(rows):
    rr=r0+i
    Gc=f"{Ccp}{rr}"; Dc=f"{Cbuy}{rr}"; Tc=f"{Ctb}{rr}"; Mc=f"{Cmg}{rr}"
    ws.cell(rr,1,r["brand"]); ws.cell(rr,2,r["keyword"]); ws.cell(rr,3,r["model"])
    ws.cell(rr,4,r["danawaBuy"])
    du=ws.cell(rr,5,r["danawaUrl"]);  du.hyperlink=(r["danawaUrl"] or None); du.font=link
    cu=ws.cell(rr,6,r["coupangUrl"]); cu.hyperlink=(r["coupangUrl"] or None); cu.font=link
    ws.cell(rr,7,r["coupangPrice"])
    ws.cell(rr,8,f'=IF({Gc}="","",ROUND({Gc}-{Dc}-{Gc}*{FEE}-{Tc},0))')
    ws.cell(rr,9,f'=IF(OR({Gc}="",{Gc}=0),"",ROUND({Mc}/{Gc}*100,1))')
    ws.cell(rr,10,f'=IF({Gc}="","위너가필요",IF({Mc}>=10,"✅마진OK","❌미달"))')
    ws.cell(rr,11,round(r["danawaBuy"]/(1-FEE)) if r["danawaBuy"] else None)
    ws.cell(rr,12,0)
    ws.cell(rr,13,r["domesticSeller"]); ws.cell(rr,14,r["sourcingNote"]); ws.cell(rr,15,r["cap"])
    ws.cell(rr,16,r["danawaSellers"]); ws.cell(rr,17,r["reviews"]); ws.cell(rr,18,r["winnerShip"])
    ws.cell(rr,19,r["sellerCountNew"]); ws.cell(rr,20," + ".join(sorted(r["src"])))
    ws.cell(rr,21,r["signal"]); ws.cell(rr,22,r["risk"])
    for j in range(1,NC+1):
        c=ws.cell(rr,j); c.border=bd
        if j not in (5,6): c.font=cf
        c.fill=PatternFill("solid",fgColor=FILL[COLS[j-1][2]])
        if j in (4,7,8,11): c.alignment=rgt; c.number_format='#,##0'
        elif j==9: c.alignment=rgt; c.number_format='0.0'
        elif j in (3,5,6,13,14,20,21,22): c.alignment=lft
        else: c.alignment=ctr
    ws.row_dimensions[rr].height=28
last=r0+len(rows)-1

ws.conditional_formatting.add(f"{Cmg}{r0}:{Cmg}{last}",CellIsRule(operator="greaterThanOrEqual",formula=["10"],fill=PatternFill("solid",fgColor="C6EFCE"),font=Font(color="006100",bold=True,name="맑은 고딕")))
ws.conditional_formatting.add(f"{Cmg}{r0}:{Cmg}{last}",CellIsRule(operator="lessThan",formula=["10"],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(color="9C0006",name="맑은 고딕")))
ws.conditional_formatting.add(f"{Cjd}{r0}:{Cjd}{last}",FormulaRule(formula=[f'{Cjd}{r0}="✅마진OK"'],fill=PatternFill("solid",fgColor="63BE7B"),font=Font(bold=True,color="FFFFFF",name="맑은 고딕")))
ws.conditional_formatting.add(f"{Cjd}{r0}:{Cjd}{last}",FormulaRule(formula=[f'{Cjd}{r0}="❌미달"'],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(color="9C0006",name="맑은 고딕")))
ws.conditional_formatting.add(f"{Cws}{r0}:{Cws}{last}",FormulaRule(formula=[f'ISNUMBER(SEARCH("로켓",{Cws}{r0}))'],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(color="9C0006",bold=True,name="맑은 고딕")))

# ---- Sheet2 ----
g=wb.create_sheet("사용법_및_판정"); g.column_dimensions["A"].width=3; g.column_dimensions["B"].width=112
lines=[("H1","통합 시트 사용법 · 쿠팡 수수료 11.88% 기준"),("P",""),
 ("H2","1. 앞쪽 열 (한눈에)"),
 ("P","브랜드 · 키워드 · 모델명 · 다나와 매입가(국내최저) · 다나와 URL · 쿠팡 URL · 쿠팡 판매가 → 그 뒤 예상마진/마진율/마진≥10원 판정."),
 ("P","A~G열은 화면 고정되어 오른쪽 스크롤해도 항상 보입니다."),
 ("H2","2. 마진 계산 (사용자 지정: 쿠팡 수수료 11.88%)"),
 ("P","예상마진 = 쿠팡 판매가 − 다나와 매입가 − 쿠팡 판매가 × 11.88% − 택배비(기본 0)."),
 ("P","'마진≥10원?': 예상마진 ≥10원이면 ✅마진OK(초록), 미만이면 ❌미달(빨강), 쿠팡가 없으면 '위너가필요'."),
 ("P","손익분기 쿠팡가 = 매입가 / 0.8812. 쿠팡 위너가가 이 값보다 높아야 마진이 남습니다."),
 ("H2","3. 쿠팡 판매가 (파란 셀)"),
 ("P","일부 모델은 업로드 파일에서 실측된 쿠팡가가 이미 들어 있음(파란 배경). 빈칸은 쿠팡 로그인으로 위너가 확인해 입력 → 마진 자동계산."),
 ("P","쿠팡은 봇차단으로 자동취득 불가라 미확인분은 비워둠(추정 안 함)."),
 ("H2","4. 소싱 규칙 반영"),
 ("P","· 해외배송/해외직구 = 제외(판매자점수 리스크). · 품절 = 제외. · 카드전용가 = 기본가 기준(비고)."),
 ("P","· 소싱 비고 '쿠팡이최저' = 다나와 최저가가 쿠팡 자신 → 매입-재판매 마진 없음, 소싱 부적합."),
 ("P","· 위너 배송형태 로켓형(빨강) = 일반 판매자배송 경쟁 불리 → 제외 대상."),
 ("H2","5. 데이터 출처 / 보존"),
 ("P","'데이터출처' 열로 각 행 근거 표시. 원본 업로드 실측표는 '쿠팡실측_업로드원본' 탭에 그대로 보존."),
]
rr=1
for k,t in lines:
    c=g.cell(rr,2,t)
    if k=="H1": c.font=Font(bold=True,size=13,color="1F3864",name="맑은 고딕")
    elif k=="H2": c.font=Font(bold=True,size=11,color="1F3864",name="맑은 고딕")
    else: c.font=Font(size=10,name="맑은 고딕")
    c.alignment=lft; rr+=1

# ---- Sheet3 업로드 원본 ----
u=wb.create_sheet("쿠팡실측_업로드원본")
uh=["브랜드","모델(원문)","용량","쿠팡가(실측)","쿠팡리뷰수","위너배송형태","현재판매자","다나와가","상태","추천판정","핵심사유","쿠팡URL","다나와URL"]
for j,h in enumerate(uh,1):
    c=u.cell(1,j,h); c.font=hdrf; c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=ctr; c.border=bd
for i,m in enumerate(UP["cand"],2):
    vals=[m.get("brand"),m.get("model"),m.get("cap"),num(m.get("coupangPrice")),m.get("reviews"),
          m.get("winnerShip"),m.get("seller"),num(m.get("danawa")),m.get("status"),m.get("verdict"),
          str(m.get("reason",""))[:80],m.get("coupangUrl"),m.get("danawaUrl")]
    for j,v in enumerate(vals,1):
        c=u.cell(i,j,v); c.font=cf; c.border=bd; c.alignment=lft if j in (2,7,11,12,13) else ctr
        if j in (4,8): c.number_format='#,##0'
for j,w in enumerate([10,26,6,11,9,14,14,11,8,18,34,30,30],1): u.column_dimensions[get_column_letter(j)].width=w
u.freeze_panes="A2"

wb.save(OUT)
print("SAVED:",OUT)
print("통합 행수:",len(rows))
print("실측 쿠팡가 보유:",sum(1 for r in rows if r["coupangPrice"]))
print("국내최저 소싱처 보유:",sum(1 for r in rows if r["domesticSeller"]))
PY_DONE=1
