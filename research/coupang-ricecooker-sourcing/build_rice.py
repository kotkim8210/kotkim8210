#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""밥솥 쿠팡 아이템위너 소싱 마진시트 (제습기 시트와 동일 방법론·포맷)"""
import json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

HERE=os.path.dirname(os.path.abspath(__file__))
FEE=0.1188
def J(f): return json.load(open(os.path.join(HERE,f),encoding="utf-8"))
w1=J("wave1.json")+J("wave1b.json")
w2={r["model"]:r for r in J("wave2.json")+J("wave2b.json")}
CPURL={"CRP-DHP0610FW":("https://www.coupang.com/vp/products/8605638836",263990,"판매중·판매자배송(폴센트 7/10 실측). 동일 모델·색상 확인"),
       "KRC-BS9906":("https://www.coupang.com/vp/products/8700969099",28430,"판매중·판매자배송(폴센트 7/10 실측). 역대최저 27,300")}

def cap_kw(name):
    m=re.search(r"(\d+)인용", name or "")
    return f"{m.group(1)}인용 밥솥" if m else "밥솥"

rows=[]
for m in w1:
    p=w2.get(m["model"],{})
    cp=p.get("coupangPrice") or 0
    buy=p.get("cheapestNonCoupangPrice") or m.get("danawaLowest") or 0
    url_cp, cp_real, cp_note = CPURL.get(m["model"], ("",None,""))
    if cp_real: cp=cp_real
    margin=round(cp-buy-cp*FEE) if (cp and buy) else None
    rocket=bool(p.get("coupangRocket"))
    note=p.get("note","") or ""
    rows.append({
      "brand":m["brand"],"kw":f"{cap_kw(m['name'])} / {m['brand']} 밥솥","model":m["model"],"name":m["name"],
      "buy":buy,"buySeller":p.get("cheapestNonCoupangSeller",""),
      "durl":f"https://prod.danawa.com/info/?pcode={m['pcode']}" if m.get("pcode") else f"https://search.danawa.com/dsearch.php?query={m['model']}",
      "curl":url_cp or (f"https://www.coupang.com/np/search?q={m['model']}"),
      "cp":cp if cp else None,"margin":margin,"rocket":rocket,
      "second":p.get("secondNonCoupang",""),"sellers":m.get("sellerCount",""),
      "listed":p.get("coupangListed"),
      "note":(cp_note or "")+((" | " if cp_note and note else "")+note if note else ""),
    })
# 정렬: 정상 양수 → 얇은 양수 → 음수(마진순)
def sk(r):
    m=r["margin"]
    grp = 0 if (m is not None and m>=10 and not r["rocket"]) else (1 if m is not None else 2)
    return (grp, -(m if m is not None else -9e9))
rows.sort(key=sk)

NAVY="1F3864"; hdrf=Font(bold=True,color="FFFFFF",size=10,name="맑은 고딕")
cf=Font(size=10,name="맑은 고딕"); thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
rgt=Alignment(horizontal="right",vertical="center")
link=Font(size=9,color="0563C1",underline="single",name="맑은 고딕")

wb=Workbook(); ws=wb.active; ws.title="통합_소싱마진_밥솥"
COLS=[("브랜드",9),("키워드",17),("모델명",19),("다나와 매입가\n(비쿠팡 국내최저)",14),("다나와 URL",26),("쿠팡 URL",26),
      ("쿠팡 판매가\n(다나와등재/실측)",13),("예상마진(원)\n수수료11.88%",13),("마진율\n(%)",8),("마진≥10원?",10),
      ("손익분기\n쿠팡가",11),("매입 판매처",13),("2위 매입처",16),("다나와\n판매처수",8),("위너\n배송추정",11),("비고/검증",34)]
NC=len(COLS)
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NC)
b=ws.cell(1,1,"밥솥 쿠팡 아이템위너 소싱 — 기준일 2026-07-10(KST) | 쿠팡 판매가=다나와 등재 쿠팡가(실조회)+폴센트 실측 | 예상마진=쿠팡가−매입가−쿠팡가×11.88% | 품절·로켓·쿠팡최저는 후보 제외")
b.font=Font(bold=True,color="FFFFFF",size=10,name="맑은 고딕"); b.fill=PatternFill("solid",fgColor=NAVY); b.alignment=lft
ws.row_dimensions[1].height=26
for j,(nm,w) in enumerate(COLS,1):
    c=ws.cell(2,j,nm); c.font=hdrf; c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=ctr; c.border=bd
    ws.column_dimensions[get_column_letter(j)].width=w
ws.row_dimensions[2].height=30
ws.freeze_panes="H3"
r0=3
for i,r in enumerate(rows):
    rr=r0+i
    cpc=f"G{rr}"; buyc=f"D{rr}"
    ws.cell(rr,1,r["brand"]); ws.cell(rr,2,r["kw"]); ws.cell(rr,3,r["model"])
    ws.cell(rr,4,r["buy"] or None)
    du=ws.cell(rr,5,r["durl"]); du.hyperlink=r["durl"]; du.font=link
    cu=ws.cell(rr,6,r["curl"]); cu.hyperlink=r["curl"]; cu.font=link
    ws.cell(rr,7,r["cp"])
    ws.cell(rr,8,f'=IF(OR({cpc}="",{buyc}=""),"",ROUND({cpc}-{buyc}-{cpc}*{FEE},0))')
    ws.cell(rr,9,f'=IF(OR({cpc}="",{cpc}=0),"",ROUND(H{rr}/{cpc}*100,1))')
    ws.cell(rr,10,f'=IF({cpc}="","쿠팡가없음",IF(H{rr}>=10,"✅마진OK","❌미달"))')
    ws.cell(rr,11,round((r["buy"] or 0)/(1-FEE)) if r["buy"] else None)
    ws.cell(rr,12,r["buySeller"]); ws.cell(rr,13,r["second"]); ws.cell(rr,14,r["sellers"])
    ws.cell(rr,15,"🚀로켓표기" if r["rocket"] else ("미등재" if r["listed"] is False else "일반추정"))
    ws.cell(rr,16,r["note"])
    for j in range(1,NC+1):
        c=ws.cell(rr,j); c.border=bd
        if j not in (5,6): c.font=cf
        if j in (4,7,8,11): c.alignment=rgt; c.number_format='#,##0'
        elif j==9: c.alignment=rgt; c.number_format='0.0'
        elif j in (3,5,6,12,13,16): c.alignment=lft
        else: c.alignment=ctr
    ws.row_dimensions[rr].height=26
last=r0+len(rows)-1
ws.conditional_formatting.add(f"H{r0}:H{last}",CellIsRule(operator="greaterThanOrEqual",formula=["10"],fill=PatternFill("solid",fgColor="C6EFCE"),font=Font(color="006100",bold=True,name="맑은 고딕")))
ws.conditional_formatting.add(f"H{r0}:H{last}",CellIsRule(operator="lessThan",formula=["10"],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(color="9C0006",name="맑은 고딕")))
ws.conditional_formatting.add(f"O{r0}:O{last}",FormulaRule(formula=[f'ISNUMBER(SEARCH("로켓",O{r0}))'],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(color="9C0006",bold=True,name="맑은 고딕")))

# 마진후보 탭
c2=wb.create_sheet("오늘기준_마진후보")
ch=["순위","브랜드","모델명","매입가(비쿠팡최저)","매입처","쿠팡 판매가","예상마진(11.88%공제)","마진율(%)","다나와 URL","쿠팡 URL","검증/주의"]
for j,h in enumerate(ch,1):
    c=c2.cell(1,j,h); c.font=hdrf; c.fill=PatternFill("solid",fgColor="375623"); c.alignment=ctr; c.border=bd
for j,w in enumerate([5,9,18,13,12,11,14,9,28,28,42],1): c2.column_dimensions[get_column_letter(j)].width=w
ri=2
for r in rows:
    m=r["margin"]
    if m is None or m<10 or r["rocket"]: continue
    vals=[ri-1,r["brand"],r["model"],r["buy"],r["buySeller"],r["cp"],m,round(m/r["cp"]*100,1),r["durl"],r["curl"],r["note"]]
    for j,v in enumerate(vals,1):
        c=c2.cell(ri,j,v); c.font=cf; c.border=bd
        if j in (4,6,7): c.number_format='#,##0'; c.alignment=rgt
        elif j==8: c.number_format='0.0'; c.alignment=rgt
        elif j in (9,10): c.alignment=lft; c.hyperlink=v or None; c.font=link
        elif j in (3,5,11): c.alignment=lft
        else: c.alignment=ctr
        if j==7: c.fill=PatternFill("solid",fgColor="C6EFCE")
    c2.row_dimensions[ri].height=26; ri+=1
if ri==2: c2.cell(2,1,"(마진 ≥10원 후보 없음)")
c2.freeze_panes="A2"

# 가이드
g=wb.create_sheet("사용법_및_결론"); g.column_dimensions["A"].width=3; g.column_dimensions["B"].width=110
lines=[
 ("H1","밥솥 소싱 조사 — 방법·결론 (2026-07-10 KST)"),("P",""),
 ("H2","방법 (제습기 조사와 동일 파이프라인)"),
 ("P","① 다나와에서 쿠쿠/쿠첸/키친아트 등 45종 수집(모델번호·최저가·판매처수·pcode) ② 각 상품페이지에서 '쿠팡 등재가'와 '비쿠팡 국내최저가(해외·품절·카드전용 제외)'를 오늘 기준 실조회 ③ 예상마진=쿠팡가−매입가−쿠팡가×11.88% ④ 양수 후보는 폴센트로 쿠팡 판매상태·배송형태 실측"),
 ("H2","냉정한 결론"),
 ("P","· 45종 중 마진 ≥10원은 단 2종. 밥솥은 모델당 다나와 판매처가 300~600곳(제습기의 2~3배)인 극단적 출혈경쟁 카테고리 + 쿠쿠몰/쿠첸 공식몰이 가격 통제 → 쿠팡 등재가가 비쿠팡 최저가와 거의 같거나 오히려 더 쌈(역마진)."),
 ("P","· 유일한 실질 후보: 쿠쿠 CRP-DHP0610FW (6인용 IH). 쿠팡 263,990원 판매중·판매자배송(폴센트 실측, 동일 모델·색상 확인) vs 옥션 매입 210,350원 → 마진 +22,278원(8.4%). 단, 스프레드 5만원대가 유지될지는 옥션 재고·쿠팡 위너 변동에 달림 — 소량(1~3개) 테스트만 권장."),
 ("P","· KRC-BS9906(키친아트 소형)은 +333원으로 사실상 무의미."),
 ("P","· 로켓 표기 6종(CRP-LHTR1010FGIM, CRH-TWS1010W, CRP-EHS0320FSM, CRH-TWS0610W, CRT-RPS0691BW, CJE-CD0302, CRT-TWK0370CY)은 윙 경쟁 불가 → 제외."),
 ("H2","주의"),
 ("P","· 쿠팡 등재가는 다나와 연동 기준이라 실시간과 다를 수 있음 — 등록 전 쿠팡 URL 클릭해 위너가·판매자수·리뷰 확인 필수."),
 ("P","· 매입 전 옥션/G마켓 등 매입처 직접 진입해 결제단계 최종가 확인(링크 경유가와 직진입가 상이 가능)."),
]
rr=1
for k,t in lines:
    c=g.cell(rr,2,t)
    if k=="H1": c.font=Font(bold=True,size=13,color="1F3864",name="맑은 고딕")
    elif k=="H2": c.font=Font(bold=True,size=11,color="1F3864",name="맑은 고딕")
    else: c.font=Font(size=10,name="맑은 고딕")
    c.alignment=lft; rr+=1

OUT=os.path.join(HERE,"쿠팡_밥솥_소싱마진시트.xlsx")
wb.save(OUT)
print("SAVED:",OUT," rows:",len(rows))
pos=[r for r in rows if r["margin"] is not None and r["margin"]>=10 and not r["rocket"]]
print("마진후보:",[(r['model'],r['margin']) for r in pos])
