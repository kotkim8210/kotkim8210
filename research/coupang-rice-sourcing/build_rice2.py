#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""쌀(수향미 등 브랜드쌀) 쿠팡 아이템위너 소싱 마진시트.
새 규칙: 재고≤10개 소싱처 제외 / 구매제한(1일1개 등) 판매자 제외 / 에누리·네이버 최저가 교차확인."""
import json, os, glob, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

HERE=os.path.dirname(os.path.abspath(__file__))
FEE=0.1188
scans=[]
for f in sorted(glob.glob(os.path.join(HERE,"scan*.json"))):
    scans+=json.load(open(f,encoding="utf-8"))
try: DEEP=json.load(open(os.path.join(HERE,"deep.json"),encoding="utf-8"))
except Exception: DEEP={}
META=json.load(open(os.path.join(HERE,"meta.json"),encoding="utf-8"))  # pcode->brand/name/sellers/weight

def num(x):
    try: return float(str(x).replace(",","").strip()) or None
    except: return None

rows=[]
for s in scans:
    pc=str(s.get("pcode",""))
    m=META.get(pc,{})
    item=s.get("item") if isinstance(s.get("item"),str) else str(s.get("item"))
    deep=DEEP.get(pc,{})
    cp=num(s.get("coupangPrice"))
    dnBuy=num(s.get("cheapestNonCoupangPrice"))
    enuri=num(deep.get("enuriLowest")); naver=num(deep.get("naverLowest"))
    # 최종 매입가 = 3개 소스 중 최저 (확인된 것만)
    cands=[(dnBuy,"다나와:"+str(s.get("cheapestNonCoupangSeller") or "")),(enuri,"에누리:"+str(deep.get("enuriSeller") or "")),(naver,"네이버")]
    cands=[(p,l) for p,l in cands if p]
    buy,buyLabel=(min(cands,key=lambda x:x[0]) if cands else (None,""))
    if deep.get("coupangCurrent"): cp=num(deep["coupangCurrent"])
    margin=round(cp-buy-cp*FEE) if (cp and buy) else None
    rocket=bool(s.get("coupangRocket"))
    limitNote=str(deep.get("limitOrStockNote") or "")
    soldout=bool(re.search(r"품절",limitNote+str(deep.get("saleStatus") or "")))
    limited=bool(re.search(r"1일\s*1개|1인\s*1개|구매제한|재고\s*[0-9]\b|한정",limitNote))
    excluded=soldout or limited
    note_parts=[]
    if deep.get("saleStatus"): note_parts.append("판매상태:"+str(deep["saleStatus"]))
    if limitNote: note_parts.append("⚠"+limitNote[:50])
    if s.get("soldOutNote"): note_parts.append(str(s.get("soldOutNote"))[:45])
    if deep.get("note"): note_parts.append(str(deep["note"])[:45])
    rows.append({
      "brand":m.get("brand",""),"item":item,"kw":m.get("kw","브랜드쌀"),
      "buy":buy,"buyLabel":buyLabel,"dnBuy":dnBuy,"enuri":enuri,"naver":naver,
      "durl":f"https://prod.danawa.com/info/?pcode={pc}",
      "curl":deep.get("coupangUrl") or f"https://www.coupang.com/np/search?q={item.split('(')[0].strip()}",
      "naverUrl":deep.get("naverUrl",""),"bestBuyUrl":deep.get("bestBuyUrl",""),
      "cp":cp,"margin":margin,"rocket":rocket,"listed":s.get("coupangListed"),
      "sellers":m.get("sellers",""),"second":str(s.get("secondNonCoupang") or ""),
      "excluded":excluded,"soldout":soldout,"note":" | ".join(note_parts),
    })
def sk(r):
    m=r["margin"]
    grp=0 if (m is not None and m>=10 and not r["rocket"] and not r["excluded"]) else (1 if m is not None else 2)
    return (grp, -(m if m is not None else -9e9))
rows.sort(key=sk)

NAVY="1F3864"; hdrf=Font(bold=True,color="FFFFFF",size=10,name="맑은 고딕")
cf=Font(size=10,name="맑은 고딕"); thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
rgt=Alignment(horizontal="right",vertical="center")
link=Font(size=9,color="0563C1",underline="single",name="맑은 고딕")

wb=Workbook(); ws=wb.active; ws.title="통합_소싱마진_쌀"
COLS=[("브랜드/생산",12),("키워드",15),("상품명(구성 주의)",26),("최종 매입가\n(3소스 최저)",12),("다나와 URL",24),("쿠팡 URL",24),
      ("쿠팡 판매가\n(등재/실측)",12),("예상마진(원)\n11.88%공제",12),("마진율\n(%)",7),("판정",10),
      ("매입가 출처",16),("다나와\n비쿠팡최저",10),("에누리\n최저",10),("네이버\n최저",10),("네이버 URL",22),
      ("다나와\n판매처수",8),("위너\n배송추정",10),("[매입전 필수확인]\n재고>10·구매제한 없음",20),("비고/검증",30)]
NC=len(COLS)
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NC)
b=ws.cell(1,1,"쌀(수향미·브랜드쌀) 쿠팡 소싱 — 기준일 2026-07-10(KST) | 매입가=다나와·에누리·네이버 3소스 최저 | 마진=쿠팡가−매입가−쿠팡가×11.88% | 제외규칙: 품절·로켓·재고≤10개·구매제한(1일1개) 판매자 | 쌀은 도정일자·구성(10kg×1 vs ×2) 반드시 일치 확인")
b.font=Font(bold=True,color="FFFFFF",size=10,name="맑은 고딕"); b.fill=PatternFill("solid",fgColor=NAVY); b.alignment=lft
ws.row_dimensions[1].height=30
for j,(nm,w) in enumerate(COLS,1):
    c=ws.cell(2,j,nm); c.font=hdrf; c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=ctr; c.border=bd
    ws.column_dimensions[get_column_letter(j)].width=w
ws.row_dimensions[2].height=32
ws.freeze_panes="H3"
r0=3
for i,r in enumerate(rows):
    rr=r0+i
    ws.cell(rr,1,r["brand"]); ws.cell(rr,2,r["kw"]); ws.cell(rr,3,r["item"])
    ws.cell(rr,4,r["buy"])
    du=ws.cell(rr,5,r["durl"]); du.hyperlink=r["durl"]; du.font=link
    cu=ws.cell(rr,6,r["curl"]); cu.hyperlink=r["curl"]; cu.font=link
    ws.cell(rr,7,r["cp"])
    ws.cell(rr,8,f'=IF(OR(G{rr}="",D{rr}=""),"",ROUND(G{rr}-D{rr}-G{rr}*{FEE},0))')
    ws.cell(rr,9,f'=IF(OR(G{rr}="",G{rr}=0),"",ROUND(H{rr}/G{rr}*100,1))')
    verdict=("제외(품절)" if r.get("soldout") else "제외(구매제한/재고)") if r["excluded"] else ("제외(로켓)" if r["rocket"] else ("미등재" if r["listed"] is False else ""))
    ws.cell(rr,10,verdict if verdict else f'=IF(G{rr}="","쿠팡가없음",IF(H{rr}>=10,"✅마진OK","❌미달"))')
    ws.cell(rr,11,r["buyLabel"]); ws.cell(rr,12,r["dnBuy"]); ws.cell(rr,13,r["enuri"]); ws.cell(rr,14,r["naver"])
    nu=ws.cell(rr,15,r["naverUrl"]);
    if r["naverUrl"]: nu.hyperlink=r["naverUrl"]; nu.font=link
    ws.cell(rr,16,r["sellers"])
    ws.cell(rr,17,"🚀로켓표기" if r["rocket"] else ("미등재" if r["listed"] is False else "일반추정"))
    ws.cell(rr,18,"⚠확인필요(봇 접근불가 항목): 매입처 페이지에서 잔여재고 10개 초과 & 1일 구매제한 없음 확인 후 매입")
    ws.cell(rr,19,r["note"])
    for j in range(1,NC+1):
        c=ws.cell(rr,j); c.border=bd
        if j not in (5,6,15): c.font=cf
        if j in (4,7,8,12,13,14): c.alignment=rgt; c.number_format='#,##0'
        elif j==9: c.alignment=rgt; c.number_format='0.0'
        elif j in (3,5,6,11,15,18,19): c.alignment=lft
        else: c.alignment=ctr
    ws.row_dimensions[rr].height=28
last=r0+len(rows)-1
ws.conditional_formatting.add(f"H{r0}:H{last}",CellIsRule(operator="greaterThanOrEqual",formula=["10"],fill=PatternFill("solid",fgColor="C6EFCE"),font=Font(color="006100",bold=True,name="맑은 고딕")))
ws.conditional_formatting.add(f"H{r0}:H{last}",CellIsRule(operator="lessThan",formula=["10"],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(color="9C0006",name="맑은 고딕")))
ws.conditional_formatting.add(f"Q{r0}:Q{last}",FormulaRule(formula=[f'ISNUMBER(SEARCH("로켓",Q{r0}))'],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(color="9C0006",bold=True,name="맑은 고딕")))
ws.conditional_formatting.add(f"J{r0}:J{last}",FormulaRule(formula=[f'ISNUMBER(SEARCH("제외",J{r0}))'],fill=PatternFill("solid",fgColor="FFC7CE"),font=Font(color="9C0006",name="맑은 고딕")))

c2=wb.create_sheet("오늘기준_마진후보")
ch=["순위","브랜드","상품명","최종 매입가","매입처(출처)","쿠팡 판매가","예상마진(11.88%)","마진율(%)","다나와URL","쿠팡URL","네이버URL","매입전 필수확인/주의"]
for j,h in enumerate(ch,1):
    c=c2.cell(1,j,h); c.font=hdrf; c.fill=PatternFill("solid",fgColor="375623"); c.alignment=ctr; c.border=bd
for j,w in enumerate([5,10,26,11,16,11,13,9,24,24,24,40],1): c2.column_dimensions[get_column_letter(j)].width=w
ri=2
for r in rows:
    m=r["margin"]
    if m is None or m<10 or r["rocket"] or r["excluded"]: continue
    vals=[ri-1,r["brand"],r["item"],r["buy"],r["buyLabel"],r["cp"],m,round(m/r["cp"]*100,1),r["durl"],r["curl"],r["naverUrl"],
          ("재고>10·구매제한없음 확인 필수 | "+r["note"])[:120]]
    for j,v in enumerate(vals,1):
        c=c2.cell(ri,j,v); c.font=cf; c.border=bd
        if j in (4,6,7): c.number_format='#,##0'; c.alignment=rgt
        elif j==8: c.number_format='0.0'; c.alignment=rgt
        elif j in (9,10,11): c.alignment=lft; c.hyperlink=v or None; c.font=link
        elif j in (3,5,12): c.alignment=lft
        else: c.alignment=ctr
        if j==7: c.fill=PatternFill("solid",fgColor="C6EFCE")
    c2.row_dimensions[ri].height=26; ri+=1
if ri==2: c2.cell(2,1,"(마진 ≥10원 후보 없음)")
c2.freeze_panes="A2"

g=wb.create_sheet("사용법_및_결론"); g.column_dimensions["A"].width=3; g.column_dimensions["B"].width=112
lines=[
 ("H1","쌀(수향미·브랜드쌀) 소싱 조사 — 방법·결론 (2026-07-10 KST)"),("P",""),
 ("H2","방법"),
 ("P","① 다나와에서 수향미·이천쌀(알찬미/해들미)·철원오대·신동진·고시히카리 28종 수집 ② 각 상품페이지에서 쿠팡 등재가+비쿠팡 국내최저 실조회 ③ 주요 후보는 에누리·네이버 최저가 교차확인(요청 반영) ④ 매입가=3소스 중 최저 ⑤ 마진=쿠팡가−매입가−쿠팡가×11.88%"),
 ("H2","신규 제외 규칙 (요청 반영)"),
 ("P","· 잔여재고 ≤10개 소싱처 제외 · 1일 1개 등 구매제한 판매자 제외(쿠팡 2개 주문시 당일 발주 불가) — 단, 재고/구매제한은 옥션·G마켓·11번가가 봇을 차단해 자동확인 불가 → '매입전 필수확인' 열에 체크리스트로 명시, 매입 직전 판매자 페이지에서 직접 확인 필수."),
 ("H2","쌀 카테고리 특이사항"),
 ("P","· 도정일자 = 사실상 유통기한. 같은 pcode라도 판매처별 도정일이 다름 — 오래된 도정미 매입시 반품 폭탄. 매입 전 도정일 문의 필수."),
 ("P","· 구성 혼동 주의: 다나와 페이지 중 '10kg×2' 구성이 섞여 있음(진광미·포천해들·영광신동진). 쿠팡 매칭 시 단위 확인."),
 ("P","· 쿠팡 쌀은 로켓프레시 비중 높음 — 위너 배송형태 확인 필수."),
 ("H2","결론"),
 ("P","· 28종 중 마진 ≥10원 3종뿐, 실질 후보는 이천쌀 알찬미 특등급 10kg 1종(+2,197원, 4.8%). 쌀은 kg당 마진이 원래 얇은 박리다매 카테고리 + 농협/RPC가 전 채널 동시 공급이라 스프레드가 거의 없음."),
 ("P","· 수향미는 브랜드 인지도는 최고지만 판매처 812곳 초경쟁이라 쿠팡 등재가(41,600)가 비쿠팡최저(39,760)에 붙어 있어 역마진."),
 ("P","· 쌀로 수익 내려면 소매 차익거래가 아니라 산지 RPC/농협 직거래 계약(20kg 단위 도매)으로 원가를 낮추는 구조가 필요."),
]
rr=1
for k,t in lines:
    c=g.cell(rr,2,t)
    if k=="H1": c.font=Font(bold=True,size=13,color="1F3864",name="맑은 고딕")
    elif k=="H2": c.font=Font(bold=True,size=11,color="1F3864",name="맑은 고딕")
    else: c.font=Font(size=10,name="맑은 고딕")
    c.alignment=lft; rr+=1

OUT=os.path.join(HERE,"쿠팡_쌀_소싱마진시트.xlsx")
wb.save(OUT)
print("SAVED:",OUT," rows:",len(rows))
pos=[r for r in rows if r["margin"] is not None and r["margin"]>=10 and not r["rocket"] and not r["excluded"]]
print("후보:",[(r['item'],r['margin']) for r in pos])
