#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""매입처 직접비교 시트 — 제습기·밥솥 생존 후보 9종 × 5사이트(에누리/네이버/G마켓/옥션/하이마트).
다나와 링크 경유가와 직진입가가 다른 문제 대응: 사이트별 직접 진입 URL + 확인가격."""
import json, os, urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE=os.path.dirname(os.path.abspath(__file__))
DATA=json.load(open(os.path.join(HERE,"direct.json"),encoding="utf-8"))  # list of {model, enuriDetailUrl, sites{...}, limitNote, note}
CTX=json.load(open(os.path.join(HERE,"context.json"),encoding="utf-8"))  # model -> {cat,brand,name,coupangPrice,coupangUrl,buyRef}

def q(s): return urllib.parse.quote(str(s))
def search_urls(model):
    return {
      "에누리": f"https://www.enuri.com/search.jsp?keyword={q(model)}",
      "네이버": f"https://search.shopping.naver.com/search/all?query={q(model)}",
      "G마켓": f"https://www.gmarket.co.kr/n/search?keyword={q(model)}",
      "옥션": f"https://browse.auction.co.kr/search?keyword={q(model)}",
      "하이마트": f"https://www.google.com/search?q=site:e-himart.co.kr+{q(model)}",
    }
SITE_ORDER=["에누리","네이버","G마켓","옥션","하이마트"]
FEE=0.1188

NAVY="1F3864"; hdrf=Font(bold=True,color="FFFFFF",size=10,name="맑은 고딕")
cf=Font(size=10,name="맑은 고딕"); thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
rgt=Alignment(horizontal="right",vertical="center")
link=Font(size=9,color="0563C1",underline="single",name="맑은 고딕")
GREY=PatternFill("solid",fgColor="F2F2F2"); MODELFILL=PatternFill("solid",fgColor="DDEBF7")

wb=Workbook(); ws=wb.active; ws.title="매입처_직접비교"
COLS=[("구분",8),("브랜드",10),("모델명",18),("사이트",9),("확인가격(원)\n※직진입가와 다를 수 있음",14),
      ("상품 직링크 (있으면 이걸로)",40),("직접 검색 진입 URL (링크 안타고 시작)",40),
      ("쿠팡 판매가\n(참고)",11),("마진 계산\n(이 가격 매입시)",12),("재고/구매제한·비고",30)]
NC=len(COLS)
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=NC)
b=ws.cell(1,1,"매입처 직접비교 — 기준일 2026-07-10(KST) | 다나와 경유가≠직진입 결제가 문제 대응: 각 사이트 직접 들어가 장바구니→결제단계 최종가로 비교하세요 | 확인가격은 조사시점 표시가(카드전용 제외), 실제 결제가는 직진입으로 재확인 | 매입 전: 재고>10개 & 1일 구매제한 없음 확인")
b.font=Font(bold=True,color="FFFFFF",size=10,name="맑은 고딕"); b.fill=PatternFill("solid",fgColor=NAVY); b.alignment=lft
ws.row_dimensions[1].height=30
for j,(nm,w) in enumerate(COLS,1):
    c=ws.cell(2,j,nm); c.font=hdrf; c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=ctr; c.border=bd
    ws.column_dimensions[get_column_letter(j)].width=w
ws.row_dimensions[2].height=32
ws.freeze_panes="A3"

r=3
for d in DATA:
    model=d["model"]; ctx=CTX.get(model,{})
    su=search_urls(model)
    cp=ctx.get("coupangPrice")
    # 모델 헤더행 (쿠팡 정보)
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=NC)
    hc=ws.cell(r,1,f"▶ {ctx.get('cat','')} | {ctx.get('brand','')} {model} — {ctx.get('name','')}   |   쿠팡 판매가 {cp:,}원 ({ctx.get('coupangUrl','')})   |   손익분기 매입가(11.88%공제 후 마진0): {round(cp*(1-FEE)):,}원 이하로 매입해야 마진" if cp else f"▶ {ctx.get('cat','')} | {ctx.get('brand','')} {model}")
    hc.font=Font(bold=True,size=10,color="1F3864",name="맑은 고딕"); hc.fill=MODELFILL; hc.alignment=lft
    ws.row_dimensions[r].height=22
    r+=1
    for site in SITE_ORDER:
        sd=(d.get("sites") or {}).get(site) or {}
        price=sd.get("price") or None
        purl=sd.get("url") or ""
        if site=="에누리" and not purl: purl=d.get("enuriDetailUrl","")
        margin=round(cp-price-cp*FEE) if (cp and price) else None
        vals=[ctx.get("cat",""),ctx.get("brand",""),model,site,price,purl,su[site],cp,margin,
              (d.get("limitNote","") if site=="옥션" else "") or sd.get("note","")]
        for j,v in enumerate(vals,1):
            c=ws.cell(r,j,v); c.font=cf; c.border=bd
            if j in (5,8,9): c.alignment=rgt; c.number_format='#,##0'
            elif j in (6,7):
                c.alignment=lft
                if v: c.hyperlink=v; c.font=link
            elif j in (3,10): c.alignment=lft
            else: c.alignment=ctr
            if margin is not None and j==9:
                c.fill=PatternFill("solid",fgColor=("C6EFCE" if margin>=10 else "FFC7CE"))
        ws.row_dimensions[r].height=22
        r+=1
    # 모델별 공통 비고
    if d.get("note"):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=NC)
        nc2=ws.cell(r,1,"   비고: "+str(d["note"])[:220]); nc2.font=Font(size=9,italic=True,color="595959",name="맑은 고딕"); nc2.alignment=lft; nc2.fill=GREY
        r+=1
    r+=1  # 빈 행

g=wb.create_sheet("사용법"); g.column_dimensions["A"].width=3; g.column_dimensions["B"].width=110
lines=[
 ("H1","매입처 직접비교 시트 사용법"),
 ("P",""),
 ("P","1. 각 모델 블록의 '손익분기 매입가'보다 싸게 사야 마진이 남습니다 (쿠팡 판매가 × 0.8812 기준)."),
 ("P","2. '상품 직링크'가 있으면 그걸로, 없으면 '직접 검색 진입 URL'로 각 사이트에 링크 안타고 들어가서 검색 → 장바구니 → 결제단계 최종가 확인. 다나와/에누리 경유가와 다르면 직진입가 기준으로 판단."),
 ("P","3. 네이버·G마켓·옥션·하이마트는 봇 차단이라 확인가격이 비어있는 칸이 많습니다 — URL은 유효하니 직접 확인하세요."),
 ("P","4. 매입 전 필수: 잔여재고 10개 초과 + 1일 구매제한 없음 + (밥솥) 정확 모델·색상 일치 + 카드전용가 아닌 기본가."),
 ("P","5. '마진 계산' 열은 해당 확인가격으로 매입했을 때 기준입니다. 직진입가가 더 싸면 마진은 더 커집니다."),
]
rr=1
for k,t in lines:
    c=g.cell(rr,2,t)
    c.font=Font(bold=True,size=13,color="1F3864",name="맑은 고딕") if k=="H1" else Font(size=10,name="맑은 고딕")
    c.alignment=lft; rr+=1

OUT=os.path.join(HERE,"매입처_직접비교_제습기_밥솥.xlsx")
wb.save(OUT)
print("SAVED:",OUT," models:",len(DATA))
