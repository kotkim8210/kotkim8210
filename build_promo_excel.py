# -*- coding: utf-8 -*-
"""서울/수도권 홍보처 추천 목록 엑셀 생성.

시트: 안내 / 건강원·건강식품점 / 요양병원(수도권) / 한방병원(수도권) /
      건강검진센터 / 로컬푸드직매장
열: A 업체명(병원명) B 지역 C 전화번호 D 이메일 E 팩스 F 홈페이지 주소
    G 우선순위(구매력) H 비고
"""

import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUTPUT = "홍보처_추천목록.xlsx"
METRO = ("서울", "경기", "인천")

# 구매력 상위 지역 키워드 (공백 제거 후 부분일치)
TIER1_KEYS = [
    "강남구", "서초구", "송파구", "용산구", "분당", "판교", "과천", "하남",
    "수지", "영통", "광교", "동안구", "평촌", "동탄", "연수구", "송도", "위례",
]
TIER2_KEYS = [
    "마포", "성동구", "양천", "영등포", "여의도", "종로", "서울중구", "강동구",
    "광진", "동작", "강서구", "동대문", "고양", "일산", "부천", "김포", "광명",
    "군포", "산본", "의왕", "구리", "남양주", "다산", "파주", "운정", "수정구",
    "중원구", "수원", "기흥", "화성", "용인", "안양", "인천서구", "청라",
    "부평", "남동구", "성남",
]


def tier(region: str) -> str:
    compact = region.replace(" ", "")
    if any(k in compact for k in TIER1_KEYS):
        return "1순위"
    if any(k in compact for k in TIER2_KEYS):
        return "2순위"
    return "3순위"


def sido_order(region: str) -> int:
    for idx, s in enumerate(METRO):
        if region.startswith(s):
            return idx
    return len(METRO)


def sort_rows(rows: list[dict]) -> list[dict]:
    return sorted(
        rows,
        key=lambda r: (r["tier"], sido_order(r["region"]), r["region"], r["name"]),
    )


def load_json(path: str) -> list[dict]:
    if not os.path.exists(path):
        print(f"!! 파일 없음: {path}")
        return []
    with open(path, encoding="utf-8") as fp:
        return json.load(fp)


def norm(rows: list[dict]) -> list[dict]:
    out = []
    for r in rows:
        region = (r.get("region") or "").strip()
        if not region.startswith(METRO):
            continue
        out.append(
            {
                "name": (r.get("name") or "").strip(),
                "region": region,
                "phone": (r.get("phone") or "").strip(),
                "email": (r.get("email") or "").strip(),
                "fax": (r.get("fax") or "").strip(),
                "homepage": (r.get("homepage") or "").strip(),
                "note": (r.get("note") or "").strip(),
                "tier": tier(region),
            }
        )
    return out


HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TIER_FILL = {
    "1순위": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "2순위": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
}


def write_sheet(wb: Workbook, title: str, name_header: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title)
    columns = [name_header, "지역", "전화번호", "이메일", "팩스", "홈페이지 주소", "우선순위", "비고"]
    for col, text in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    keys = ["name", "region", "phone", "email", "fax", "homepage", "tier", "note"]
    for r_idx, row in enumerate(sort_rows(rows), start=2):
        for c_idx, key in enumerate(keys, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row[key])
        fill = TIER_FILL.get(row["tier"])
        if fill:
            ws.cell(row=r_idx, column=7).fill = fill

    widths = {"A": 38, "B": 18, "C": 16, "D": 22, "E": 14, "F": 42, "G": 10, "H": 38}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows) + 1}"


GUIDE = [
    ["식품애착 기름 홍보처 추천 목록 (서울/수도권, 구매력 우선)"],
    [""],
    ["시트 구성 (공략 우선순위 순)"],
    ["1. 건강원·건강식품점", "보양식·중장년 타깃과 가장 잘 맞는 1차 공략 채널"],
    ["2. 요양병원(수도권)", "보호자·간병인 대상 주변 상권 공략용 기준점 (직접 입점보다 인근 상권 활용)"],
    ["3. 한방병원(수도권)", "보양·체질 메시지와 연결되는 제휴 후보"],
    ["4. 건강검진센터", "건강 관심층 밀집 — 대기공간 비치·제휴 제안용"],
    ["5. 로컬푸드직매장", "원물 신뢰도·선물 수요 — 입점(위탁) 제안용"],
    [""],
    ["우선순위(G열) 기준 — 구매력 상위 지역"],
    ["1순위", "서울 강남·서초·송파·용산 / 성남 분당(판교)·과천·하남·용인 수지·수원 영통(광교)·안양 동안(평촌)·화성 동탄 / 인천 연수(송도) 등"],
    ["2순위", "서울 마포·성동·양천·영등포·종로 등 / 고양(일산)·부천·김포·수원·용인 기흥 등 / 인천 서구(청라)·부평·남동"],
    ["3순위", "그 외 수도권 지역"],
    [""],
    ["데이터 출처"],
    ["요양병원·한방병원", "건강보험심사평가원(HIRA) 병원·약국 찾기 (공식 데이터, 전수)"],
    ["건강원·검진센터·로컬푸드", "웹 검색 교차 확인 (대표 업체 선별 수집)"],
    [""],
    ["참고", "이메일·팩스는 공공 데이터에 공개되지 않아 대부분 빈칸 — 전화/홈페이지 문의폼 활용 권장"],
]


def write_guide(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "안내"
    for r_idx, row in enumerate(GUIDE, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif len(row) == 1 and val:
                cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100


def main() -> None:
    yoyang = [r for r in load_json("data/yoyang_hospitals_full.json") if r["region"].startswith(METRO)]
    for r in yoyang:
        r["note"] = ""
        r["tier"] = tier(r["region"])
    hanbang = load_json("data/hanbang_metro.json")
    for r in hanbang:
        r["note"] = ""
        r["tier"] = tier(r["region"])

    gunkang = norm(load_json("data/health_stores.json"))
    checkup = norm(load_json("data/checkup_centers.json"))
    localfood = norm(load_json("data/localfood_stores.json"))

    wb = Workbook()
    write_guide(wb)
    write_sheet(wb, "건강원·건강식품점", "업체명", gunkang)
    write_sheet(wb, "요양병원(수도권)", "병원명", yoyang)
    write_sheet(wb, "한방병원(수도권)", "병원명", hanbang)
    write_sheet(wb, "건강검진센터", "기관명", checkup)
    write_sheet(wb, "로컬푸드직매장", "매장명", localfood)
    wb.save(OUTPUT)

    print(f"저장: {OUTPUT}")
    for label, rows in [
        ("건강원", gunkang), ("요양병원", yoyang), ("한방병원", hanbang),
        ("검진센터", checkup), ("로컬푸드", localfood),
    ]:
        t1 = sum(1 for r in rows if r["tier"] == "1순위")
        t2 = sum(1 for r in rows if r["tier"] == "2순위")
        print(f"  {label}: {len(rows)}건 (1순위 {t1} / 2순위 {t2} / 3순위 {len(rows)-t1-t2})")


if __name__ == "__main__":
    main()
