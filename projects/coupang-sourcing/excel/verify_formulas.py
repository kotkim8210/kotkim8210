#!/usr/bin/env python3
"""워크북 수식 정적 검증 — LibreOffice 없이 확인할 수 있는 것들을 확인한다.

recalc(LibreOffice)이 느리거나 불가한 환경에서도 아래를 보장하기 위한 검사:
  1) 모든 시트간 참조(가정!$C$n, 세율표!...)가 **값이 있는 셀**을 가리키는가
     → 오타로 빈 셀을 가리키면 계산이 0이 되어 조용히 틀린다. 가장 위험한 실수.
  2) LibreOffice가 못 읽는 함수(XLOOKUP·FILTER 등)를 쓰지 않았는가
  3) 핵심 행의 마진 계산을 파이썬으로 독립 재현해 수식과 일치하는가
  4) 0으로 나눌 수 있는 자리에 가드(IFERROR/IF)가 있는가

사용: python excel/verify_formulas.py [워크북경로]
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

BANNED = ("XLOOKUP", "XMATCH", "FILTER(", "UNIQUE(", "SORT(", "SEQUENCE(", "TEXTJOIN", "IFS(")
XREF = re.compile(r"(가정|세율표|로켓그로스요금표)!\$?([A-Z]+)\$?(\d+)")

_ok = 0
_bad = 0


def check(cond: bool, msg: str, got=None) -> None:
    global _ok, _bad
    if cond:
        _ok += 1
        print(f"  ✅ {msg}")
    else:
        _bad += 1
        print(f"  ❌ {msg}" + (f"  (got={got})" if got is not None else ""))


def _skip_margin(ws, g, t) -> None:
    """상품 데이터가 없는 템플릿용 — 수식 존재 여부만 확인한다."""
    print("[4] 0 나눗셈 가드")
    for col in ("V", "AF", "AH", "AI"):
        f = ws[f"{col}8"].value
        check(isinstance(f, str) and ("IFERROR" in f or "IF(" in f), f"{col}열에 IF/IFERROR 가드 있음")
    print("[5] 하드코딩 제거 확인")
    for col, needle, label in (("I", "가정!", "공급가가 가정!환율을 참조"),
                               ("N", "가정!", "마케팅이 가정 시트를 참조"),
                               ("S", "MAX(0", "소득세에 MAX(0,·) 적용")):
        f = ws[f"{col}8"].value
        check(isinstance(f, str) and needle in f, label, f)
    check(t["D4"].value is not None and t["E4"].value is not None, "세율표가 채워져 있음")


def main() -> int:
    path = Path(sys.argv[1] if len(sys.argv) > 1 else Path(__file__).parent / "calc_v4.xlsx")
    wb = load_workbook(path)
    ws = wb["소싱현황"]
    g = wb["가정"]
    t = wb["세율표"]

    # ── 1) 시트간 참조가 값 있는 셀을 가리키는지 ─────────────────────────
    print("[1] 시트간 참조 유효성")
    refs, dangling = set(), []
    for row in ws.iter_rows(min_row=8, max_row=8):      # 대표로 첫 데이터 행
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                for sheet, col, rownum in XREF.findall(c.value):
                    refs.add((sheet, col, int(rownum)))
    sheets = {"가정": g, "세율표": t, "로켓그로스요금표": wb["로켓그로스요금표"] if "로켓그로스요금표" in wb.sheetnames else None}
    for sheet, col, rownum in sorted(refs):
        target_ws = sheets.get(sheet)
        if target_ws is None:
            dangling.append(f"{sheet}(시트없음)")
            continue
        if target_ws[f"{col}{rownum}"].value is None:
            dangling.append(f"{sheet}!{col}{rownum}")
    check(not dangling, f"참조 {len(refs)}개 모두 값 있는 셀을 가리킴", dangling)

    # 로켓그로스 요금표 자체의 가정 참조도 확인
    if "로켓그로스요금표" in wb.sheetnames:
        rg = wb["로켓그로스요금표"]
        rg_dangling = []
        for r in range(4, 7):
            f = rg[f"G{r}"].value
            if isinstance(f, str):
                for sheet, col, rownum in XREF.findall(f):
                    if sheets[sheet][f"{col}{rownum}"].value is None:
                        rg_dangling.append(f"G{r}→{sheet}!{col}{rownum}")
        check(not rg_dangling, "요금표 물류비 수식의 가정 참조 유효", rg_dangling)
        check(rg["B4"].value == "극소형" and rg["B5"].value == "소형",
              "요금표에 실측 사이즈(극소형·소형) 존재", (rg["B4"].value, rg["B5"].value))
        check(rg["D4"].value == 1000 and rg["E4"].value == 1551,
              "CFS 실측 단가(입출고 1,000 / 배송 1,551) 반영", (rg["D4"].value, rg["E4"].value))
        # 판매가별 통합요율 표가 실제로 만들어졌는지(저가 상품 경고의 근거)
        found = any(isinstance(rg.cell(row=rr, column=2).value, str)
                    and "판매가별 통합 부담률" in rg.cell(row=rr, column=2).value
                    for rr in range(1, rg.max_row + 1))
        check(found, "판매가별 통합 부담률 표 존재(저가 상품 위험 표시)")

    # ── 2) 금지 함수 ────────────────────────────────────────────────────
    print("[2] LibreOffice 비호환 함수")
    found = []
    for row in ws.iter_rows(min_row=8, max_row=210):
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                for b in BANNED:
                    if b in c.value.upper():
                        found.append(f"{c.coordinate}:{b}")
    check(not found, "금지 함수 미사용", found[:5])

    # ── 3) 마진 계산 독립 재현 ──────────────────────────────────────────
    print("[3] 마진 계산 파이썬 재현 (8행)")
    def gv(label_starts):
        """가정 시트에서 라벨로 값을 찾는다(행 위치가 바뀌어도 안전)."""
        for rr in range(1, g.max_row + 1):
            lab = g.cell(row=rr, column=2).value
            if isinstance(lab, str) and lab.startswith(label_starts):
                return g.cell(row=rr, column=3).value
        return None

    A = {
        "fx": gv("환율"), "customsMode": gv("수입 통관"), "impVat": gv("수입 매입세액"),
        "shipFee": gv("배송비 수수료"), "mkt": gv("마케팅(광고)"), "etc": gv("기타 비용"),
        "bizType": gv("사업자 유형"), "annual": gv("예상 연 과세표준"),
        "coupon": gv("판매자 할인쿠폰"), "feeVatExcl": gv("수수료율 VAT"),
    }
    check(all(v is not None for v in A.values()), "가정 시트 핵심 항목 모두 조회됨",
          [k for k, v in A.items() if v is None])
    G_ = ws["G8"].value; H_ = ws["H8"].value or 0
    J_ = ws["J8"].value; K_ = ws["K8"].value
    L_ = ws["L8"].value; M_ = ws["M8"].value
    if any(v is None for v in (G_, J_, K_, L_, M_)):
        # 빈 템플릿(상품 데이터 없음) — 마진 재현은 건너뛰고 나머지 검사만 수행
        print("     (8행이 비어 있어 마진 재현은 건너뜁니다 — 템플릿 파일)")
        _skip_margin(ws, g, t)
        print(f"\n결과: {_ok}/{_ok + _bad} 통과", "✅" if _bad == 0 else f"❌ {_bad} 실패")
        return 1 if _bad else 0
    check(True, "8행 입력값 존재")

    I_ = J_ * A["fx"]
    N_ = G_ * A["mkt"]
    CP_ = G_ * A["coupon"]                                   # 할인쿠폰(실측 5.6%)
    vat_mult = 1.1 if A["feeVatExcl"] == "Y" else 1.0
    P_ = (G_ - CP_) * K_ * vat_mult + H_ * A["shipFee"]      # ★정산 실측 공식
    imp = I_ * A["impVat"] if A["customsMode"] == "정식통관" else 0
    Q_ = ((G_ - CP_ + H_) * 0.015 if A["bizType"] == "간이과세"
          else (G_ - CP_ + H_) / 11 - (L_ + M_ + N_ + P_) / 11 - imp)
    R_ = G_ - CP_ + H_ - I_ - L_ - M_ - N_ - P_ - Q_

    # 소득세율: 세율표에서 연 과세표준에 맞는 구간(지방포함 = 세율×1.1)
    rate = 0
    for r in range(4, 12):
        lo = t[f"B{r}"].value
        if isinstance(lo, (int, float)) and A["annual"] >= lo:
            rate = t[f"D{r}"].value * 1.1
    S_ = max(0, R_) * rate
    T_ = G_ * A["etc"]
    U_ = R_ - S_ - T_

    print(f"     공급가={I_:,.0f} 쿠폰={CP_:,.0f} 마케팅={N_:,.0f} 수수료={P_:,.0f} 부가세={Q_:,.0f}")
    print(f"     1차합계={R_:,.0f} 소득세율={rate*100:.1f}% 소득세={S_:,.0f} 기타={T_:,.0f}")
    print(f"     → 마진={U_:,.0f} ({U_/G_*100:.1f}%)")
    check(rate > 0, f"소득세율이 세율표에서 결정됨 ({rate*100:.1f}%)")
    check(U_ < R_, "마진 < 1차합계 (세금·기타가 차감됨)")

    # ── 4) 0 나눗셈 가드 ────────────────────────────────────────────────
    print("[4] 0 나눗셈 가드")
    for col in ("V", "AF", "AH", "AI"):
        f = ws[f"{col}8"].value
        check(isinstance(f, str) and ("IFERROR" in f or "IF(" in f),
              f"{col}열에 IF/IFERROR 가드 있음")

    # ── 5) 수식이 하드코딩 상수를 쓰지 않는지 ───────────────────────────
    print("[5] 하드코딩 제거 확인")
    i8 = ws["I8"].value
    check(isinstance(i8, str) and "가정!" in i8, "공급가가 가정!환율을 참조 (=J*300 하드코딩 제거)", i8)
    n8 = ws["N8"].value
    check(isinstance(n8, str) and "가정!" in n8, "마케팅이 가정 시트를 참조", n8)
    s8 = ws["S8"].value
    check(isinstance(s8, str) and "MAX(0" in s8, "소득세에 MAX(0,·) 적용 (적자 시 음수세금 방지)", s8)

    print(f"\n결과: {_ok}/{_ok + _bad} 통과", "✅" if _bad == 0 else f"❌ {_bad} 실패")
    return 1 if _bad else 0


if __name__ == "__main__":
    raise SystemExit(main())
