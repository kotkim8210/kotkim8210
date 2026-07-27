#!/usr/bin/env python3
"""크롤러 결과 → 마진계산기 엑셀의 '소싱 입력'(연두칸) 자동 채우기.

이게 '자동 소싱기'의 자동 부분이다.
사람이 채워야 하는 연두칸(AA~AE: 경쟁사 최저가·리뷰중앙·이길만한 경쟁%·광고침투율·월검색량)을
쿠팡 크롤링 + 네이버 검색광고 API 결과로 채워 넣는다. 나머지는 엑셀 수식이 알아서 판정한다.

사용:
    # C열(상품명)을 검색어로 삼아 전 행 조사
    python excel/fill_sourcing.py 마진계산기_V4_일반사업자.xlsx

    # 특정 행만 (8~20행)
    python excel/fill_sourcing.py 파일.xlsx --rows 8-20

    # 키워드를 따로 지정 (상품명이 검색어로 부적합할 때)
    python excel/fill_sourcing.py 파일.xlsx --keyword-col AJ

주의: 값을 쓰고 나면 수식 캐시가 비므로 재계산이 필요하다.
      python /root/.claude/skills/xlsx/scripts/recalc.py <파일>
      (또는 엑셀에서 그냥 열면 자동 재계산된다)
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from sourcing import competition, naver_ads

DATA_START = 8


def parse_rows(spec: str | None, ws, name_col: str) -> list[int]:
    if spec:
        lo, _, hi = spec.partition("-")
        return list(range(int(lo), int(hi or lo) + 1))
    rows = []
    for r in range(DATA_START, ws.max_row + 1):
        if ws[f"{name_col}{r}"].value:
            rows.append(r)
    return rows


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="크롤러 결과로 엑셀 소싱칸 자동 채우기")
    ap.add_argument("workbook", help="마진계산기 xlsx 경로")
    ap.add_argument("--rows", help="처리할 행 범위 (예: 8-30). 생략 시 상품명 있는 전 행")
    ap.add_argument("--name-col", default="C", help="검색어로 쓸 열 (기본 C=상품명)")
    ap.add_argument("--keyword-col", help="검색어 전용 열이 따로 있으면 지정")
    ap.add_argument("--sheet", default="소싱현황")
    ap.add_argument("--pages", type=int, default=1)
    ap.add_argument("--limit", type=int, default=40)
    ap.add_argument("--no-crawl", action="store_true", help="검색량만 채우고 쿠팡 크롤링은 생략")
    args = ap.parse_args(argv)

    path = Path(args.workbook)
    wb = load_workbook(path)          # 수식 보존 모드 (data_only=False)
    ws = wb[args.sheet]
    rows = parse_rows(args.rows, ws, args.name_col)
    print(f"대상 {len(rows)}행", file=sys.stderr)

    crawl = None
    if not args.no_crawl:
        try:
            crawler_path = Path(__file__).resolve().parent.parent.parent / "coupang-naver-crawler"
            sys.path.insert(0, str(crawler_path))
            from crawler import coupang
            from crawler.common import FetchOptions
            crawl = (coupang, FetchOptions(engine="stealthy"))
        except Exception as exc:  # noqa: BLE001
            print(f"크롤러 로드 실패 → 검색량만 채웁니다: {exc}", file=sys.stderr)

    filled = 0
    for r in rows:
        kw = ws[f"{args.keyword_col}{r}"].value if args.keyword_col else ws[f"{args.name_col}{r}"].value
        if not kw:
            continue
        kw = str(kw).strip()
        print(f"  [{r}] {kw}", file=sys.stderr)

        # 월 검색량 (네이버 검색광고 API — 키 없으면 None)
        stats = naver_ads.keyword_stats(kw)
        if stats:
            ws[f"AE{r}"] = stats["total"]

        # 쿠팡 경쟁강도
        if crawl:
            coupang_mod, opts = crawl
            try:
                items = coupang_mod.search(kw, pages=args.pages, opts=opts, limit=args.limit)
                m = competition.analyze(items)
                if not m.get("empty"):
                    ws[f"AA{r}"] = m["price_min"]
                    ws[f"AB{r}"] = m["review_median"]
                    ws[f"AC{r}"] = m["beatable_ratio"]
                    ws[f"AD{r}"] = m["ad_ratio"]
                    print(f"       → {competition.summarize(m)}", file=sys.stderr)
            except Exception as exc:  # noqa: BLE001
                print(f"       크롤링 실패: {str(exc)[:90]}", file=sys.stderr)
        filled += 1

    wb.save(path)
    print(f"\n✅ {filled}행 갱신 → {path}", file=sys.stderr)
    print("※ 판정/점수를 보려면 엑셀에서 파일을 열거나(자동 재계산), recalc.py를 실행하세요.",
          file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
