#!/usr/bin/env python3
"""마진계산기 V4 (일반사업자 기준) + 자동 소싱기 워크북 생성기.

원본: 사용자의 '마진계산기 V3.0'(법인 기준, 소싱현황 단일 시트)
바뀐 점:
  1) 법인세(1차합계×16% 고정) → **종합소득세 누진세율 + 지방소득세** 자동 적용
  2) 부가세를 '마진×10%' 근사 → **매출세액 − 매입세액** 정밀 계산으로 (간이과세 옵션 포함)
  3) 환율·수수료·마케팅% 등 하드코딩(=J*300, G*15%) → **가정 시트 참조**로 분리
  4) 소싱 판정 열 추가(경쟁사가·리뷰·검색량·광고침투율 → 손익분기가·기회점수·판정)
     → 빈칸을 채우면 그대로 자동 소싱기가 된다.

사용: python excel/build_calculator.py [출력경로]
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else HERE / "마진계산기_V4_일반사업자.xlsx"
PRODUCTS = HERE / "products.json"

# ── 서식 (원본 관례 계승: LG Smart UI SemiBold, 회계형 숫자) ──────────────
FONT = "LG Smart UI SemiBold"
F_TITLE = Font(name=FONT, size=14, bold=True, color="FFFFFF")
F_HEAD = Font(name=FONT, size=9, bold=True)
F_BODY = Font(name=FONT, size=9)
F_IN = Font(name=FONT, size=9, color="0000FF")      # 파랑 = 직접 입력
F_CALC = Font(name=FONT, size=9, color="000000")    # 검정 = 수식
F_LINK = Font(name=FONT, size=9, color="008000")    # 초록 = 다른 시트 참조
F_NOTE = Font(name=FONT, size=8, italic=True, color="666666")

FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_HEAD = PatternFill("solid", fgColor="D9E1F2")
FILL_IN = PatternFill("solid", fgColor="FFF2CC")     # 노랑 = 채워야 할 칸
FILL_SRC = PatternFill("solid", fgColor="E2EFDA")    # 연두 = 소싱 입력
FILL_OUT = PatternFill("solid", fgColor="FCE4D6")    # 주황 = 최종 판정

NUM = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@'
CNY = '0.00_);[Red]\\(0.00\\)'
PCT = '0.0%'
SCORE = '0.0'

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

DATA_START = 8          # 상품 데이터 시작 행 (원본과 동일)
N_ROWS = 200            # 빈 행 포함 총 행 수(소싱 후보를 계속 추가할 수 있게)


def a(sheet: str, cell: str) -> str:
    """다른 시트 참조 (시트명에 공백 없으므로 따옴표 불필요)."""
    return f"{sheet}!{cell}"


# =========================================================================== #
# 1) 가정 시트
# =========================================================================== #
def build_assumptions(ws) -> dict:
    ws.title = "가정"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 52

    ws.merge_cells("B2:D2")
    ws["B2"] = "가정 (여기만 고치면 전체 계산이 바뀝니다)"
    ws["B2"].font = F_TITLE
    ws["B2"].fill = FILL_TITLE
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    ref: dict[str, str] = {}
    r = 4

    def section(title: str):
        nonlocal r
        ws.cell(row=r, column=2, value=title).font = Font(name=FONT, size=10, bold=True)
        ws.cell(row=r, column=2).fill = FILL_HEAD
        ws.cell(row=r, column=3).fill = FILL_HEAD
        ws.cell(row=r, column=4).fill = FILL_HEAD
        r += 1

    def item(key: str, label: str, value, note: str, fmt: str | None = None, is_input=True):
        nonlocal r
        ws.cell(row=r, column=2, value=label).font = F_BODY
        c = ws.cell(row=r, column=3, value=value)
        c.font = F_IN if is_input else F_CALC
        if is_input:
            c.fill = FILL_IN
        c.border = BORDER
        if fmt:
            c.number_format = fmt
        ws.cell(row=r, column=4, value=note).font = F_NOTE
        ref[key] = f"$C${r}"
        r += 1

    section("① 사업자 / 세금")
    item("bizType", "사업자 유형", "일반과세",
         "일반과세 / 간이과세 중 입력. 간이과세는 부가세를 매출×1.5%로 근사")
    item("annualIncome", "예상 연 과세표준(원)", 50000000,
         "이 값으로 종합소득세 한계세율이 자동 결정됩니다(세율표 시트)", NUM)
    item("incomeTaxRate", "→ 적용 소득세율(지방포함)", None,
         "자동 계산: 종합소득세 + 지방소득세 10%", PCT, is_input=False)
    item("vatRate", "부가가치세율", 0.10, "기본 10%", PCT)

    section("② 수입 / 원가  ★LCL 인보이스 4건 실측 반영")
    item("fx", "환율 CNY→KRW (체감)", 300,
         "★실측 검증: LCL 3건 역산 결과 248/278/349원 → 평균 291원. 300원 가정은 타당(보수적)", NUM)
    item("customsMode", "수입 통관 방식", "정식통관",
         "정식통관 / 구매대행. 정식통관이면 수입부가세를 매입세액으로 공제")
    item("importVatCredit", "수입 매입세액 공제율", 0.10,
         "정식통관 시 공급가 대비 공제율(≈10%). 구매대행·직송금이면 공제 0", PCT)
    item("dutyRate", "관세율", 0.0,
         "★실측: 한중FTA 원산지증명서(CO) 발급 시 관세 0원(4건 중 3건). CO 미발급 1건만 관세 부과", PCT)
    item("logiFixed", "수입 고정비(건당)", 111980,
         "★실측: 통관수수료 33,000 + 도크이용료 24,200 + B/L발급 21,780 + CO발급 33,000", NUM)
    item("logiPerCbm", "수입 변동비(원/CBM)", 118000,
         "★실측: 해운운임+중국내륙+핸들링+기타 합계 ÷ CBM (3건 평균 118,000 · 오차 0.2~2%)", NUM)
    item("palletFee", "팔레트비(개당)", 38500, "★실측: 38,500원/개 (약 1.5CBM당 1개)", NUM)

    section("③ 쿠팡 수수료 / 비용  ★정산내역 683건 실측 반영")
    item("feeDefault", "서비스이용율(VAT별도)", 0.106,
         "★실측: 683건 중 652건이 10.6%, 19건 5.8%, 11건 10.8%. 판매액 대비 실질 10.86%", PCT)
    item("feeVatExcl", "수수료율 VAT 별도인가", "N",
         "★실측 공식: 판매수수료 = (판매액−할인쿠폰) × 서비스이용율 × 1.1 [683건 중 400건 표본 100% 일치]. "
         "기본 N: 이 시트의 K열은 VAT 포함 실부담률입니다(로켓그로스 30% all-in, 마켓플레이스 12%≈10.6%×1.1). K열에 서비스이용율(10.6%) 원값을 넣을 때만 Y로 바꾸세요")
    item("couponRate", "판매자 할인쿠폰율", 0.056,
         "★실측: 판매액 27,435,450 중 쿠폰 1,535,610 = 5.6%. 즉시할인+다운로드 쿠폰. "
         "쿠폰은 수수료 계산 전에 차감되고 정산액도 그만큼 줄어든다", PCT)
    item("shipFeeRate", "배송비 수수료율", 0.0363,
         "원본 V3의 3.63% 유지(배송비에 부과되는 수수료)", PCT)
    item("mktRate", "마케팅(광고) 기본(%)", 0.15,
         "원본 V3의 '=판매가×15%' 유지. ⚠️ 전량 광고판매 가정이면 과대 — 아래 광고믹스 참고", PCT)
    item("etcRate", "기타 비용(%)", 0.03, "원본 V3의 '=판매가×3%' 유지(반품·잡비 등)", PCT)
    item("returnRate", "반품율", 0.02, "참고용(마케팅·기타에 이미 반영돼 있으면 중복 주의)", PCT)

    section("④ 광고 믹스 (무광고 판매 가능성 판단)")
    item("cpc", "광고 CPC(원)", 300, "쿠팡 광고 클릭당 비용", NUM)
    item("cvr", "전환율", 0.02, "클릭 대비 구매 전환율", PCT)
    item("adCostPerSale", "→ 광고 판매 1건당 광고비", None,
         "= CPC ÷ 전환율. 전량 광고로 팔면 이 금액이 개당 광고비", NUM, is_input=False)
    item("adShare", "광고 판매 비중", 0.2,
         "전체 판매 중 광고로 팔리는 비율. 낮출수록 실질 광고비↓ (오가닉 확보가 핵심)", PCT)
    item("adBlended", "→ 개당 실질 광고비", None,
         "= 건당 광고비 × 광고 판매 비중. N열(마케팅) 대신 쓰려면 아래 스위치 ON", NUM, is_input=False)
    item("useAdMix", "마케팅에 광고믹스 사용", "N",
         "Y면 마케팅 = 위 '개당 실질 광고비', N이면 마케팅 = 판매가×마케팅%")

    section("⑤ 소싱 판정 기준")
    item("marginFloor", "마진 안전선", 0.10, "이 아래면 '위험'", PCT)
    item("marginGood", "마진 우수선", 0.30, "이 이상이면 마진점수 만점", PCT)
    item("reviewHard", "리뷰 상한(경쟁 과열)", 300, "경쟁사 리뷰 중앙값이 이 이상이면 진입점수 0", NUM)
    item("demandSat", "검색량 포화점", 50000, "월 검색량이 이 값이면 수요점수 만점", NUM)
    item("wDemand", "가중치·수요", 0.30, "합이 1이 되도록", PCT)
    item("wEntry", "가중치·진입용이", 0.30, "리뷰 낮고 이길만한 경쟁자 많을수록", PCT)
    item("wLowAd", "가중치·저광고", 0.15, "광고 침투율이 낮을수록", PCT)
    item("wMargin", "가중치·마진", 0.25, "마진율이 높을수록", PCT)
    item("goScore", "GO 기준 점수", 65, "이 이상이면 GO", SCORE)
    item("holdScore", "검토 기준 점수", 45, "이 이상이면 검토, 미만이면 SKIP", SCORE)

    section("⑥ 로켓그로스 (판매방식)")
    item("channel", "기본 판매방식", "로켓그로스",
         "로켓그로스 / 마켓플레이스. 행별로 다르면 소싱현황 AJ열에서 덮어쓰기")
    item("rgAllIn", "로켓그로스 통합요율(K열)", 0.30,
         "판매수수료+입출고비+배송비 통합 실부담률. ★검산(공식요율): 수수료 10.86% + 입출고 1,375 + 배송 2,200 = 판매가 2만원 기준 28.7% → 30% 가정 타당. 이 값을 K열에 쓸 땐 위 '수수료율 VAT 별도'를 N으로")
    item("stockDays", "예상 재고 보관일수", 45,
         "쿠팡 센터 보관 예상일. 무료보관(30일, 의류·신발·악세 45일) 초과분에 보관료 부과", NUM)
    item("freeStockDays", "무료 보관일수", 30, "카테고리에 따라 30일 또는 45일", NUM)
    item("rgPromo", "신규 프로모션 적용중", "N",
         "Y면 입출고비·배송비 0원(첫 판매 개시 후 90일간, 최대 90일 또는 누적매출 2억까지). "
         "프로모션 중이면 K열 통합요율을 판매수수료만(약 11%)으로 낮추세요")
    item("saver", "로켓그로스 세이버 적용", "N",
         "Y면 보관 60일까지 무료 + 반품 회수·재입고 무제한 무료. 90일 이후에도 270일간 0원 프로모션 중")

    section("⑦ 밀크런 (내 창고→쿠팡센터 입고)  ★정산 20건 실측")
    item("mrPerPallet", "밀크런 팔레트당 단가", 29590,
         "★실측: 26,900원(VAT별도) + 세액 2,690 = 29,590원. 3팔레트 건은 25,555원/팔레트로 소폭 할인", NUM)
    item("mrBoxPerPallet", "팔레트당 박스 수", 10,
         "★실측 평균 10.0개(249박스/25팔레트). 이 값이 클수록 개당 입고비 급감 — 6개면 4,932원/박스, "
         "14개면 2,114원/박스로 2.3배 차이", NUM)
    item("mrUnitPerBox", "박스당 상품 수", 20, "내 상품 기준 1박스에 몇 개 들어가는지", NUM)
    item("mrPerUnit", "→ 개당 밀크런비", None,
         "= 팔레트당 단가 ÷ (팔레트당 박스수 × 박스당 상품수). L열(물류비)에 포함시킬 금액", NUM, is_input=False)

    section("⑧ 목표")
    item("monthTarget", "월 목표 이익(원)", 500000, "원본 V3의 '월 목표금액' 계승", NUM)

    # 자동 계산 셀 채우기
    ws[ref["incomeTaxRate"].replace("$", "")] = (
        f"=IFERROR(INDEX(세율표!$E$4:$E$11,MATCH({ref['annualIncome']},세율표!$B$4:$B$11,1)),0)"
    )
    ws[ref["adCostPerSale"].replace("$", "")] = f"=IFERROR({ref['cpc']}/{ref['cvr']},0)"
    ws[ref["adBlended"].replace("$", "")] = f"={ref['adCostPerSale']}*{ref['adShare']}"
    ws[ref["mrPerUnit"].replace("$", "")] = (
        f"=IFERROR({ref['mrPerPallet']}/({ref['mrBoxPerPallet']}*{ref['mrUnitPerBox']}),0)"
    )

    r += 1
    ws.cell(row=r, column=2, value="색상 범례").font = Font(name=FONT, size=9, bold=True)
    r += 1
    for label, fill, note in [
        ("직접 입력", FILL_IN, "노랑 = 여기를 채우세요"),
        ("소싱 입력", FILL_SRC, "연두 = 경쟁사 조사값 (소싱현황 AA~AE열)"),
        ("판정 결과", FILL_OUT, "주황 = 자동 판정 (소싱현황 AF~AI열)"),
    ]:
        ws.cell(row=r, column=2, value=label).font = F_BODY
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=4, value=note).font = F_NOTE
        r += 1

    return ref


# =========================================================================== #
# 2) 세율표 시트
# =========================================================================== #
def build_tax_table(ws):
    ws.title = "세율표"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [3, 18, 18, 14, 14, 40]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:F2")
    ws["B2"] = "종합소득세 누진세율 (개인 일반사업자) — 지방소득세 10% 포함"
    ws["B2"].font = F_TITLE
    ws["B2"].fill = FILL_TITLE
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    for i, h in enumerate(["과세표준 하한", "과세표준 상한", "소득세율", "지방포함", "비고"], start=2):
        c = ws.cell(row=3, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    brackets = [
        (0, 14_000_000, 0.06, "1,400만원 이하"),
        (14_000_000, 50_000_000, 0.15, "1,400만 ~ 5,000만"),
        (50_000_000, 88_000_000, 0.24, "5,000만 ~ 8,800만"),
        (88_000_000, 150_000_000, 0.35, "8,800만 ~ 1.5억"),
        (150_000_000, 300_000_000, 0.38, "1.5억 ~ 3억"),
        (300_000_000, 500_000_000, 0.40, "3억 ~ 5억"),
        (500_000_000, 1_000_000_000, 0.42, "5억 ~ 10억"),
        (1_000_000_000, None, 0.45, "10억 초과"),
    ]
    for i, (lo, hi, rate, note) in enumerate(brackets):
        r = 4 + i
        ws.cell(row=r, column=2, value=lo).number_format = NUM
        ws.cell(row=r, column=3, value=hi if hi else "이상").number_format = NUM
        ws.cell(row=r, column=4, value=rate).number_format = PCT
        ws.cell(row=r, column=5, value=f"=D{r}*1.1").number_format = PCT
        ws.cell(row=r, column=6, value=note).font = F_NOTE
        for c in range(2, 7):
            ws.cell(row=r, column=c).font = F_BODY
            ws.cell(row=r, column=c).border = BORDER

    # MATCH가 참조하는 '지방포함' 열을 D로 쓰기 위해 열 위치 정리:
    # B=하한, C=상한, D=세율, E=지방포함 → 가정 시트는 D열(세율) 대신 E열(지방포함)을 봐야 한다.
    # 가정 시트 수식을 E열로 맞추기 위해 여기서 D열에 지방포함값을 넣는 대신,
    # 아래 안내를 남기고 가정 시트가 INDEX(세율표!$E$4:$E$11)을 보도록 한다.

    ws.cell(row=13, column=2,
            value="※ 단위 마진에 적용하는 것은 '한계세율'입니다. 1개 더 팔 때 늘어나는 세금이므로 "
                  "누진공제는 반영하지 않습니다(원본 V3의 R×16% 방식과 동일한 취지).").font = F_NOTE
    ws.cell(row=14, column=2,
            value="※ 참고 - 법인세: 과표 2억 이하 9%(지방포함 9.9%), 2억~200억 19%(20.9%). "
                  "원본 V3의 16%는 개인 1,400만~5,000만 구간(16.5%)에 가깝습니다.").font = F_NOTE
    ws.cell(row=15, column=2,
            value="※ 4대보험·건강보험료는 미반영. 실제 수령액은 이보다 낮을 수 있습니다.").font = F_NOTE


# =========================================================================== #
# 2-b) 로켓그로스 요금표 시트
# =========================================================================== #
# 쿠팡 공식 '로켓그로스 비용 구성' 페이지 실제 고지값 (사용자 제공 캡처 기준)
RG_TIERS = [
    # (사이즈 유형, 예시, 입출고비, 배송비)
    ("Large Size 1", "전자레인지 크기", 1375, 2200),
    ("Large Size 2", "의자 크기", 1375, 4100),
    ("Extra Large Size", "서랍장 크기", 1375, 5600),
]


def build_rg_table(ws, ref: dict):
    ws.title = "로켓그로스요금표"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [3, 12, 30, 12, 12, 14, 16]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:G2")
    ws["B2"] = "로켓그로스 비용 구성 — 쿠팡 공식 고지값"
    ws["B2"].font = F_TITLE
    ws["B2"].fill = FILL_TITLE
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    for i, h in enumerate(["사이즈 유형", "예시", "입출고비", "배송비", "프로모션 적용", "실적용 물류비"], start=2):
        c = ws.cell(row=3, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    for i, (tier, basis, inout, ship) in enumerate(RG_TIERS):
        r = 4 + i
        ws.cell(row=r, column=2, value=tier)
        ws.cell(row=r, column=3, value=basis).font = F_NOTE
        ws.cell(row=r, column=4, value=inout).number_format = NUM
        ws.cell(row=r, column=5, value=ship).number_format = NUM
        ws.cell(row=r, column=6, value=f'=가정!{ref["rgPromo"]}')
        # 프로모션(첫 90일) 적용 중이면 입출고비·배송비 0원
        ws.cell(row=r, column=7,
                value=f'=IF(가정!{ref["rgPromo"]}="Y",0,D{r}+E{r})').number_format = NUM
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if cell.font.name != FONT:
                cell.font = F_BODY

    n = 4 + len(RG_TIERS) + 1
    ws.cell(row=n, column=2, value="그 외 비용 (무료 조건)").font = Font(name=FONT, size=10, bold=True)
    n += 1
    for i, h in enumerate(["항목", "기본", "프로모션(첫 90일)", "로켓세이버"], start=2):
        c = ws.cell(row=n, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
    etc = [
        ("보관비", "매 입고시 30일 무료", "최대 90일 무료", "보관 60일까지 무료"),
        ("반품 회수비", "매달 20건 무료", "무료", "무제한 무료"),
        ("반품 재입고비", "매달 20개 무료", "무료", "무제한 무료"),
        ("반출비", "매달 20개 무료", "무료", "매달 20개 무료"),
    ]
    for i, row in enumerate(etc):
        rr = n + 1 + i
        for j, v in enumerate(row):
            c = ws.cell(row=rr, column=2 + j, value=v)
            c.font = F_BODY if j == 0 else F_NOTE
            c.border = BORDER

    n2 = n + 1 + len(etc) + 1
    notes = [
        "※ 판매수수료는 '판매자배송과 동일'(카테고리별). 위 입출고비·배송비는 판매된 상품에만 부과됩니다.",
        "※ 프로모션: 첫 판매 개시 후 90일간 로켓그로스 시작비용 0원 — 최대 90일 또는 누적매출 2억원 달성 시까지.",
        "※ 로켓그로스 세이버: 90일 이후에도 270일간 0원 프로모션(보관 60일 미경과 상품 대상).",
        "※ 프로모션 신청기간 2026.07.01~07.31 · 대상: 기간 내 판매 개시한 로켓그로스 신규 판매자.",
        "※ 부가서비스: 반출 배송 90일 무료(최대 10만원), 바코드 부착·오류수정은 별도.",
        "※ 카테고리·사이즈유형·판매가에 따라 최종 비용이 결정되므로, wing의 '입출고/배송 비용 확인하기'로 상품별 실요율을 확인하세요.",
        "",
        "▶ 검산: 판매가 20,000원 · Large Size 1 기준",
        "   판매수수료 10.86%(2,172) + 입출고비 1,375 + 배송비 2,200 = 5,747원 = 판매가의 28.7%",
        "   → 소싱현황 K열의 통합요율 30% 가정이 실제와 부합합니다(프로모션 종료 후 기준).",
    ]
    for i, t in enumerate(notes):
        ws.cell(row=n2 + i, column=2, value=t).font = F_NOTE


# =========================================================================== #
# 2-c) 수입실적 시트 (LCL 인보이스 4건 실측)
# =========================================================================== #
LCL = [
    # 날짜, CARTONS, KGS, CBM, 관세, 세금(부가세/기타세금), 총액, 비고
    ("2026-01-08", 1, 178.0, 1.91, 0, 103030, 478208, "FTA CO 발급 → 관세 0"),
    ("2026-02-06", 1, 186.0, 1.61, 0, 139310, 481368, "FTA CO 발급 → 관세 0"),
    ("2026-02-12", 1, 110.0, 1.15, 48050, 52860, 359628, "CO 미발급 → 관세 48,050"),
    ("2026-03-09", 4, 638.0, 5.87, 0, 358220, 687883, "해운운임 별도 청구건(세금+부대비 위주)"),
]


def build_import_actuals(ws, ref: dict):
    ws.title = "수입실적"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJ", [3, 13, 9, 9, 9, 11, 11, 12, 12, 34]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:J2")
    ws["B2"] = "수입 실적 (FNH LOGISTICS LCL 인보이스 4건) — 가정값의 실측 근거"
    ws["B2"].font = F_TITLE
    ws["B2"].fill = FILL_TITLE
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    heads = ["통관일", "박스", "중량(kg)", "CBM", "관세", "세금", "총액", "순물류비", "원/CBM", "비고"]
    for i, h in enumerate(heads, start=2):
        c = ws.cell(row=3, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    for i, (d, ct, kg, cbm, duty, tax, total, note) in enumerate(LCL):
        r = 4 + i
        vals = [d, ct, kg, cbm, duty, tax, total]
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.font = F_BODY
            c.border = BORDER
            if j >= 4:
                c.number_format = NUM
        ws.cell(row=r, column=9, value=f"=H{r}-F{r}-G{r}").number_format = NUM   # 순물류비
        ws.cell(row=r, column=10, value=f'=IF(E{r}=0,"",I{r}/E{r})').number_format = NUM
        ws.cell(row=r, column=9).border = BORDER
        ws.cell(row=r, column=10).border = BORDER
        ws.cell(row=r, column=11, value=note).font = F_NOTE

    n = 4 + len(LCL) + 1
    lines = [
        ("물류비 구조 (회귀 결과 — 오차 0.2~2%)", True),
        ("  순물류비 = 고정비 111,980원/건  +  118,000원/CBM  +  팔레트 38,500원/개(약 1.5CBM당 1개)", False),
        ("     고정비 내역: 통관수수료 33,000 + 도크이용료 24,200 + B/L발급 21,780 + 원산지증명서(CO) 33,000", False),
        ("     변동비 내역: 해운운임 + 중국 위해항 운송료 + HANDLING CHARGE + 중국해관 인상비(1CBM당 20위안)", False),
        ("", False),
        ("관세 — 한중FTA가 핵심", True),
        ("  4건 중 3건이 관세 0원. 원산지증명서(CO, 발급비 33,000원)를 받으면 관세가 면제됩니다.", False),
        ("  CO를 안 받은 2026-02-12 건만 관세 48,050원 부과 → CO 발급비 33,000원을 아끼려다 48,050원을 냈습니다.", False),
        ("  ▶ 결론: 중국 수입은 FTA CO를 반드시 챙기세요. 가정 시트의 관세율 기본값을 0%로 둔 근거입니다.", False),
        ("", False),
        ("체감환율 300원 검증 (수입부가세로 CIF 역산 → 물품가 산출)", True),
        ("  2026-01-08: 물품가 879,410원(4,510위안) + 물류 375,178 → 랜디드 1,254,588 → 278원/위안", False),
        ("  2026-02-06: 물품가 1,265,910원(6,492위안) + 물류 342,058 → 랜디드 1,607,968 → 248원/위안", False),
        ("  2026-02-12: 물품가 389,700원(1,998위안) + 물류 258,718 + 관세 48,050 → 696,468 → 349원/위안", False),
        ("  ▶ 평균 291원/위안. 기존 '원가×300' 관례는 실측과 부합하며 약간 보수적(안전)입니다.", False),
        ("", False),
        ("소량 수입의 함정", True),
        ("  고정비 111,980원은 화물 크기와 무관하게 붙습니다. 1.15CBM 건은 고정비만 전체의 31%였습니다.", False),
        ("  ▶ CBM을 키울수록 개당 물류비가 급감합니다. 소량 테스트 수입은 원가가 구조적으로 불리합니다.", False),
    ]
    for i, (t, bold) in enumerate(lines):
        c = ws.cell(row=n + i, column=2, value=t)
        c.font = Font(name=FONT, size=9, bold=True) if bold else F_NOTE


# =========================================================================== #
# 2-d) 밀크런 실적 시트 (쿠팡 밀크런 정산 20건 실측)
# =========================================================================== #
MILKRUN = [
    # 픽업일, 요금타입, 팔레트, 박스, 발생금액, 총액(VAT포함)
    ("2026.03.11", "밀크런이용", 2, 17, 53800, 59180),
    ("2026.03.11", "취소금액", 1, 6, 32500, 35750),
    ("2026.03.12", "밀크런이용", 2, 16, 53800, 59180),
    ("2026.03.16", "밀크런이용", 1, 6, 26900, 29590),
    ("2026.03.23", "미청구", 1, 7, 0, 0),
    ("2026.03.23", "밀크런이용", 1, 7, 26900, 29590),
    ("2026.04.01", "밀크런이용", 1, 6, 26900, 29590),
    ("2026.04.01", "밀크런이용", 1, 13, 26900, 29590),
    ("2026.04.09", "밀크런이용", 1, 6, 26900, 29590),
    ("2026.04.09", "밀크런이용", 1, 12, 26900, 29590),
    ("2026.04.15", "밀크런이용", 3, 37, 76665, 84332),
    ("2026.04.22", "밀크런이용", 1, 14, 26900, 29590),
    ("2026.04.22", "밀크런이용", 1, 6, 26900, 29590),
    ("2026.04.22", "미청구", 2, 15, 0, 0),
    ("2026.04.25", "미청구", 2, 25, 0, 0),
    ("2026.04.25", "미청구", 2, 18, 0, 0),
    ("2026.04.27", "밀크런이용", 2, 25, 53800, 59180),
    ("2026.04.27", "밀크런이용", 2, 18, 53800, 59180),
    ("2026.04.29", "밀크런이용", 3, 36, 76665, 84332),
    ("2026.04.29", "밀크런이용", 2, 24, 53800, 59180),
]


def build_milkrun(ws, ref: dict):
    ws.title = "밀크런실적"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [3, 13, 12, 8, 8, 12, 14, 11]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:H2")
    ws["B2"] = "밀크런 정산 실적 (내 창고 → 쿠팡 물류센터 입고 운송) 20건"
    ws["B2"].font = F_TITLE
    ws["B2"].fill = FILL_TITLE
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    for i, h in enumerate(["픽업일", "요금타입", "팔레트", "박스", "발생금액", "총액(VAT포함)", "박스당"], start=2):
        c = ws.cell(row=3, column=i, value=h)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    for i, (d, ft, pal, box, amt, tot) in enumerate(MILKRUN):
        r = 4 + i
        for j, v in enumerate([d, ft, pal, box, amt, tot]):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.font = F_BODY
            c.border = BORDER
            if j >= 4:
                c.number_format = NUM
        c = ws.cell(row=r, column=8, value=f'=IF(E{r}=0,"",G{r}/E{r})')
        c.number_format = NUM
        c.border = BORDER

    last = 4 + len(MILKRUN) - 1
    n = last + 2
    ws.cell(row=n, column=2, value="합계 (청구건만)").font = Font(name=FONT, size=10, bold=True)
    for i, (lab, formula, fmt) in enumerate([
        ("청구 건수", f'=COUNTIFS(G4:G{last},">0")', NUM),
        ("팔레트 합계", f'=SUMIFS(D4:D{last},G4:G{last},">0")', NUM),
        ("박스 합계", f'=SUMIFS(E4:E{last},G4:G{last},">0")', NUM),
        ("총액 합계", f'=SUMIFS(G4:G{last},G4:G{last},">0")', NUM),
        ("→ 팔레트당 단가", f'=IFERROR(D{n+4}/D{n+2},"")', NUM),
        ("→ 박스당 단가", f'=IFERROR(D{n+4}/D{n+3},"")', NUM),
        ("→ 팔레트당 평균 박스", f'=IFERROR(D{n+3}/D{n+2},"")', '0.0'),
    ], start=1):
        rr = n + i
        ws.cell(row=rr, column=2, value=lab).font = F_BODY
        c = ws.cell(row=rr, column=4, value=formula)
        c.font = F_CALC
        c.number_format = fmt
        c.border = BORDER

    n2 = n + 9
    lines = [
        ("핵심: 팔레트를 꽉 채울수록 개당 입고비가 급감합니다", True),
        ("  밀크런은 팔레트 단위 과금(26,900원/팔레트, VAT 포함 29,590원)이라 박스 수와 무관합니다.", False),
        ("  실측 편차:  6박스/팔레트 → 4,932원/박스   vs   14박스/팔레트 → 2,114원/박스  (2.3배)", False),
        ("  ▶ 입고 전에 팔레트 적재율을 최대로 올리는 것이 개당 원가를 낮추는 가장 쉬운 레버입니다.", False),
        ("", False),
        ("전체 원가 사슬 (이 계산기가 반영하는 3개 구간)", True),
        ("  중국 1688 ──[LCL 해상]──▶ 내 창고 ──[밀크런]──▶ 쿠팡센터 ──[로켓그로스]──▶ 고객", False),
        ("     물품가×환율      고정 111,980 + 118,000/CBM    29,590/팔레트    입출고 1,375 + 배송 2,200~5,600", False),
        ("", False),
        ("기타 관찰", True),
        ("  · 미청구(무료) 4건 — 팔레트 7개·박스 65개분. 프로모션 또는 조건 충족 건으로 보입니다.", False),
        ("  · 3팔레트 건은 76,665원 = 25,555원/팔레트로 소폭 할인(대량 인센티브 추정).", False),
        ("  · 2026-03-11 '취소금액' 32,500원 1건 — 취소 수수료도 원가에 잡히니 주의.", False),
    ]
    for i, (t, bold) in enumerate(lines):
        c = ws.cell(row=n2 + i, column=2, value=t)
        c.font = Font(name=FONT, size=9, bold=True) if bold else F_NOTE


# =========================================================================== #
# 3) 소싱현황 시트
# =========================================================================== #
COLS = [
    ("B", 5.4, "NO"),
    ("C", 5.9, "상품명"),
    ("D", 45.4, None),          # C와 병합
    ("E", 17.4, "옵션"),
    ("F", 9.0, "수량"),
    ("G", 11.4, "판매가"),
    ("H", 12.4, "배송비"),
    ("I", 11.4, "공급가"),
    ("J", 11.4, "원가(위안)"),
    ("K", 11.4, "수수료(%)"),
    ("L", 12.4, "물류비"),
    ("M", 13.4, "포장/부자재"),
    ("N", 11.4, "마케팅"),
    ("O", 0.9, None),
    ("P", 11.4, "수수료"),
    ("Q", 11.4, "부가세"),
    ("R", 11.4, "1차 합계"),
    ("S", 11.4, "소득세"),
    ("T", 11.4, "기타"),
    ("U", 11.4, "마진"),
    ("V", 9.0, "마진율"),
    ("W", 0.9, None),
    ("X", 22.0, "공급처URL"),
    ("Y", 14.0, "공급처URL2"),
    ("Z", 22.0, "경쟁사URL"),
    ("AA", 12.0, "경쟁사 최저가"),
    ("AB", 12.0, "경쟁 리뷰중앙"),
    ("AC", 13.0, "이길만한 경쟁%"),
    ("AD", 12.0, "광고 침투율"),
    ("AE", 12.0, "월 검색량"),
    ("AF", 12.0, "손익분기가"),
    ("AG", 11.0, "가격여유"),
    ("AH", 10.0, "기회점수"),
    ("AI", 10.0, "판정"),
    ("AJ", 13.0, "판매방식"),
    ("AK", 12.0, "사이즈단계"),
    ("AL", 14.0, "권장물류비"),
    ("AM", 12.0, "할인쿠폰"),
]
SOURCING_IN = {"AA", "AB", "AC", "AD", "AE"}
SOURCING_OUT = {"AF", "AG", "AH", "AI"}
INPUT_COLS = {"C", "E", "F", "G", "H", "J", "K", "L", "M", "X", "Y", "Z", "AJ", "AK"}


def build_sourcing(ws, ref: dict, products: list[dict]):
    ws.title = "소싱현황"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4.1
    for col, w, _ in COLS:
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:N5")
    ws["B2"] = "소싱현황 관리 · 마진계산기 V4 (일반사업자)"
    ws["B2"].font = F_TITLE
    ws["B2"].fill = FILL_TITLE
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("P2:AI5")
    ws["P2"] = ("① 노랑칸(판매가·원가·수수료…)을 채우면 마진이 나오고,\n"
                "② 연두칸(경쟁사 최저가·리뷰·검색량·광고침투율)을 채우면 기회점수와 판정이 나옵니다.\n"
                "③ 세율·환율·광고믹스는 [가정] 시트에서 한 번에 바꿉니다.")
    ws["P2"].font = Font(name=FONT, size=9)
    ws["P2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["P2"].fill = PatternFill("solid", fgColor="F2F2F2")

    # 헤더 (7행)
    for col, _, head in COLS:
        if head is None:
            continue
        c = ws[f"{col}7"]
        c.value = head
        c.font = F_HEAD
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = FILL_SRC if col in SOURCING_IN else (FILL_OUT if col in SOURCING_OUT else FILL_HEAD)
    ws.row_dimensions[7].height = 28

    A = ref  # 가정 시트 참조
    G = "가정"

    for i in range(N_ROWS):
        r = DATA_START + i
        p = products[i] if i < len(products) else None
        ws.merge_cells(f"C{r}:D{r}")

        # ── 입력 열 ──────────────────────────────────────────────────
        ws[f"B{r}"] = i + 1
        if p:
            ws[f"C{r}"] = p["name"]
            ws[f"E{r}"] = p["opt"]
            ws[f"F{r}"] = p["qty"]
            ws[f"G{r}"] = p["price"]
            ws[f"H{r}"] = p["ship"] if p["ship"] is not None else 0
            ws[f"J{r}"] = p["cny"]
            ws[f"K{r}"] = p["fee"]
            ws[f"L{r}"] = p["logi"]
            ws[f"M{r}"] = p["pack"]
            ws[f"X{r}"] = p["url1"]
            ws[f"Y{r}"] = p["url2"]
            ws[f"Z{r}"] = p["comp"]
            # 공급가: 원본이 직접입력이면 그 값 유지, 수식(=J*300)이면 가정 환율 참조로 교체
            sr = p["supply_raw"]
            if isinstance(sr, str) and sr.startswith("="):
                ws[f"I{r}"] = f"=J{r}*{G}!{A['fx']}"
            elif sr is not None:
                ws[f"I{r}"] = sr
            else:
                ws[f"I{r}"] = f"=J{r}*{G}!{A['fx']}"
        else:
            ws[f"H{r}"] = 0
            ws[f"K{r}"] = (f'=IF({G}!{A["channel"]}="로켓그로스",'
                           f'{G}!{A["rgAllIn"]},{G}!{A["feeDefault"]})')
            ws[f"I{r}"] = f"=J{r}*{G}!{A['fx']}"

        # ── 계산 열 ──────────────────────────────────────────────────
        # 마케팅: 광고믹스 스위치에 따라 (판매가×마케팅%) 또는 (개당 실질 광고비)
        ws[f"N{r}"] = (f'=IF({G}!{A["useAdMix"]}="Y",{G}!{A["adBlended"]},'
                       f'G{r}*{G}!{A["mktRate"]})')
        # 할인쿠폰(AM): 실측상 판매액의 5.6%가 쿠폰으로 빠지고, 수수료도 쿠폰 차감 후 금액에 부과된다
        ws[f"AM{r}"] = f'=G{r}*{G}!{A["couponRate"]}'
        # 수수료: (판매가−쿠폰) × 요율 × [VAT별도면 1.1] + 배송비×배송수수료율
        #   ★정산내역 400건 표본 100% 일치 검증된 공식
        ws[f"P{r}"] = (f'=((G{r}-AM{r})*K{r}*IF({G}!{A["feeVatExcl"]}="Y",1.1,1))'
                       f'+(H{r}*{G}!{A["shipFeeRate"]})')
        # 부가세: 일반과세 = 매출세액 − 매입세액 / 간이과세 = 매출×1.5%
        ws[f"Q{r}"] = (
            f'=IF({G}!{A["bizType"]}="간이과세",(G{r}-AM{r}+H{r})*0.015,'
            f'(G{r}-AM{r}+H{r})/11'
            f'-((L{r}+M{r}+N{r}+P{r})/11)'
            f'-IF({G}!{A["customsMode"]}="정식통관",I{r}*{G}!{A["importVatCredit"]},0))'
        )
        ws[f"R{r}"] = f"=G{r}-AM{r}+H{r}-I{r}-L{r}-M{r}-N{r}-P{r}-Q{r}"
        ws[f"S{r}"] = f'=MAX(0,R{r})*{G}!{A["incomeTaxRate"]}'
        ws[f"T{r}"] = f'=G{r}*{G}!{A["etcRate"]}'
        ws[f"U{r}"] = f"=R{r}-S{r}-T{r}"
        ws[f"V{r}"] = f'=IF(G{r}=0,"",U{r}/G{r})'

        # ── 소싱 판정 ────────────────────────────────────────────────
        # 손익분기가: 마진 안전선을 만족하는 최소 판매가 (고정비/(1-변동비율))
        denom = (f'(1-K{r}-{G}!{A["mktRate"]}-{G}!{A["etcRate"]}'
                 f'-{G}!{A["marginFloor"]}-1/11)')
        ws[f"AF{r}"] = f'=IF({denom}<=0,"",IFERROR((I{r}+L{r}+M{r})/{denom},""))'
        # 가격여유: 경쟁사 최저가가 내 손익분기가보다 얼마나 위인가
        ws[f"AG{r}"] = f'=IF(OR(AA{r}="",AF{r}=""),"",AA{r}-AF{r})'
        # 기회점수 (가중합, 신호 없으면 그 항목 제외하고 재정규화)
        demand = f'MIN(1,LOG(AE{r}+1)/LOG({G}!{A["demandSat"]}))'
        entry = f'MAX(0,1-AB{r}/{G}!{A["reviewHard"]})*0.6+AC{r}*0.4'
        lowad = f'MAX(0,1-AD{r})'
        margin = f'MIN(1,MAX(0,(V{r}-{G}!{A["marginFloor"]})/({G}!{A["marginGood"]}-{G}!{A["marginFloor"]})))'
        wd, we, wl, wm = (f'{G}!{A["wDemand"]}', f'{G}!{A["wEntry"]}',
                          f'{G}!{A["wLowAd"]}', f'{G}!{A["wMargin"]}')
        num = (f'IF(AE{r}="",0,{wd}*{demand})+IF(AB{r}="",0,{we}*({entry}))'
               f'+IF(AD{r}="",0,{wl}*{lowad})+IF(V{r}="",0,{wm}*{margin})')
        den = (f'IF(AE{r}="",0,{wd})+IF(AB{r}="",0,{we})'
               f'+IF(AD{r}="",0,{wl})+IF(V{r}="",0,{wm})')
        ws[f"AH{r}"] = f'=IF(G{r}="","",IFERROR(({num})/({den})*100,""))'
        ws[f"AI{r}"] = (
            f'=IF(AH{r}="","",'
            f'IF(U{r}<=0,"역마진",'
            f'IF(AH{r}>={G}!{A["goScore"]},"GO",'
            f'IF(AH{r}>={G}!{A["holdScore"]},"검토","SKIP"))))'
        )
        # 판매방식(비우면 가정의 기본값) · 사이즈단계로 권장 물류비 자동 조회
        ws[f"AL{r}"] = (
            f'=IF(AK{r}="","",'
            f'IFERROR(INDEX(로켓그로스요금표!$G$4:$G$9,'
            f'MATCH(AK{r},로켓그로스요금표!$B$4:$B$9,0)),"단계명 확인"))'
        )

        # ── 서식 ────────────────────────────────────────────────────
        for col, _, head in COLS:
            if head is None:      # D(병합)·O·W(스페이서)는 건너뛴다
                continue
            c = ws[f"{col}{r}"]
            c.border = BORDER
            if col in INPUT_COLS:
                c.font = F_IN
                c.fill = FILL_IN
            elif col in SOURCING_IN:
                c.font = F_IN
                c.fill = FILL_SRC
            elif col in SOURCING_OUT:
                c.font = F_CALC
                c.fill = FILL_OUT
            else:
                c.font = F_CALC
        for col in ("G", "H", "I", "L", "M", "N", "P", "Q", "R", "S", "T", "U",
                    "AA", "AE", "AF", "AG"):
            ws[f"{col}{r}"].number_format = NUM
        ws[f"J{r}"].number_format = CNY
        for col in ("K", "V", "AC", "AD"):
            ws[f"{col}{r}"].number_format = PCT
        ws[f"AB{r}"].number_format = NUM
        ws[f"AH{r}"].number_format = SCORE
        ws[f"AI{r}"].alignment = Alignment(horizontal="center")

    last = DATA_START + N_ROWS - 1

    # 조건부 서식: 마진율 / 판정
    ws.conditional_formatting.add(
        f"V{DATA_START}:V{last}",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(name=FONT, size=9, color="9C0006", bold=True),
                   fill=PatternFill("solid", fgColor="FFC7CE")))
    ws.conditional_formatting.add(
        f"V{DATA_START}:V{last}",
        CellIsRule(operator="greaterThanOrEqual", formula=[f"가정!{ref['marginGood']}"],
                   font=Font(name=FONT, size=9, color="006100", bold=True),
                   fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(
        f"AI{DATA_START}:AI{last}",
        CellIsRule(operator="equal", formula=['"GO"'],
                   font=Font(name=FONT, size=9, color="006100", bold=True),
                   fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(
        f"AI{DATA_START}:AI{last}",
        CellIsRule(operator="equal", formula=['"역마진"'],
                   font=Font(name=FONT, size=9, color="9C0006", bold=True),
                   fill=PatternFill("solid", fgColor="FFC7CE")))
    ws.conditional_formatting.add(
        f"AI{DATA_START}:AI{last}",
        CellIsRule(operator="equal", formula=['"SKIP"'],
                   font=Font(name=FONT, size=9, color="9C6500"),
                   fill=PatternFill("solid", fgColor="FFEB9C")))

    # 하단 요약 (원본의 '월 판매량 확인' 계승)
    s = last + 2
    ws.merge_cells(f"B{s}:D{s}")
    ws[f"B{s}"] = "요약 / 목표"
    ws[f"B{s}"].font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ws[f"B{s}"].fill = FILL_TITLE

    rows = [
        ("등록 상품 수", f'=COUNTA(C{DATA_START}:C{last})', NUM),
        ("GO 판정 수", f'=COUNTIF(AI{DATA_START}:AI{last},"GO")', NUM),
        ("검토 판정 수", f'=COUNTIF(AI{DATA_START}:AI{last},"검토")', NUM),
        ("역마진 수", f'=COUNTIF(AI{DATA_START}:AI{last},"역마진")', NUM),
        ("평균 마진율", f'=IFERROR(AVERAGE(V{DATA_START}:V{last}),"")', PCT),
        ("최고 기회점수", f'=IFERROR(MAX(AH{DATA_START}:AH{last}),"")', SCORE),
        ("월 목표 이익(원)", f'=가정!{ref["monthTarget"]}', NUM),
        ("→ 최고마진 상품 기준 월 필요 판매수",
         f'=IFERROR(가정!{ref["monthTarget"]}/MAX(U{DATA_START}:U{last}),"")', NUM),
        ("→ 일 필요 판매수",
         f'=IFERROR(가정!{ref["monthTarget"]}/MAX(U{DATA_START}:U{last})/30,"")', '0.0'),
    ]
    for i, (label, formula, fmt) in enumerate(rows):
        rr = s + 1 + i
        ws.cell(row=rr, column=2, value=label).font = F_BODY
        c = ws.cell(row=rr, column=4, value=formula)
        c.font = F_CALC
        c.number_format = fmt
        c.border = BORDER

    ws.freeze_panes = f"E{DATA_START}"


# =========================================================================== #
def main() -> int:
    products = json.loads(PRODUCTS.read_text(encoding="utf-8")) if PRODUCTS.exists() else []
    wb = Workbook()
    ref = build_assumptions(wb.active)
    build_tax_table(wb.create_sheet())
    build_rg_table(wb.create_sheet(), ref)
    build_import_actuals(wb.create_sheet(), ref)
    build_milkrun(wb.create_sheet(), ref)
    build_sourcing(wb.create_sheet(), ref, products)
    wb.move_sheet("소싱현황", offset=-5)  # 소싱현황을 첫 시트로
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}  (상품 {len(products)}건, 총 {N_ROWS}행)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
