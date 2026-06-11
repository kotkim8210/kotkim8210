# -*- coding: utf-8 -*-
"""전국 요양병원 정보 수집 스크립트.

건강보험심사평가원(HIRA) 병원·약국 찾기 서비스에서 요양병원(종별코드 28)
전체 목록을 수집하고, 병원별 상세 조회로 홈페이지 주소를 보강하여
엑셀 파일로 저장한다.

출력 형식 (1행 머리글):
    A열: 병원명 / B열: 지역 / C열: 전화번호 / D열: 이메일 / E열: 팩스 / F열: 홈페이지 주소

이메일·팩스는 HIRA 등 공공 데이터에 제공되지 않아 빈칸으로 둔다.
"""

import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

SEARCH_URL = "https://www.hira.or.kr/ra/hosp/selectHospSrchKndListAjax.do"
DETAIL_URL = "https://www.hira.or.kr/ra/hosp/hospAjax.do"
CL_CD_NURSING_HOSPITAL = "28"  # HIRA 종별코드: 요양병원
OUTPUT_FILE = "요양병원_목록.xlsx"
CACHE_FILE = "/tmp/hospitals_full.json"
DETAIL_WORKERS = 3

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "X-Requested-With": "XMLHttpRequest",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "Referer": "https://www.hira.or.kr/ra/hosp/getHealthMap.do?pgmid=HIRAA030002000000",
}


def post_json(session: requests.Session, url: str, payload: dict) -> dict:
    for attempt in range(4):
        try:
            resp = session.post(url, data=payload, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            return resp.json()
        except (requests.RequestException, ValueError):
            if attempt == 3:
                raise
            time.sleep(2 ** (attempt + 1))
    raise RuntimeError("unreachable")


def collect_list(session: requests.Session) -> list[dict]:
    hospitals: dict[str, dict] = {}

    first = post_json(session, SEARCH_URL, {"pageIndex": 1, "clCd": CL_CD_NURSING_HOSPITAL, "isDown": "N"})
    pagination = first["data"]["paginationInfo"]
    total_pages = pagination["totalPageCount"]
    print(f"전체 요양병원 수: {pagination['totalRecordCount']} ({total_pages}페이지)", flush=True)

    page = 1
    while page <= total_pages:
        data = first if page == 1 else post_json(
            session, SEARCH_URL, {"pageIndex": page, "clCd": CL_CD_NURSING_HOSPITAL, "isDown": "N"}
        )
        for item in data["data"]["hospList"]:
            ykiho = item.get("ykiho") or f"{item.get('yadmNm')}|{item.get('addr')}"
            region = " ".join(
                part for part in (item.get("sidoCdNm"), item.get("sgguCdNm")) if part
            )
            hospitals[ykiho] = {
                "ykiho": ykiho,
                "name": (item.get("yadmNm") or "").strip(),
                "region": region.strip(),
                "phone": (item.get("yadmGdTelnoTxt") or "").strip(),
                "email": "",
                "fax": "",
                "homepage": "",
            }
        print(f"  목록 {page}/{total_pages} 페이지 수집 (누적 {len(hospitals)}건)", flush=True)
        page += 1
        time.sleep(0.3)

    return list(hospitals.values())


def fetch_detail(session: requests.Session, hospital: dict) -> dict:
    """병원 상세 조회로 홈페이지 주소(및 전화번호 보강)를 가져온다."""
    try:
        data = post_json(session, DETAIL_URL, {"ykiho": hospital["ykiho"]})
        info = data.get("data", {}).get("hospInfo") or {}
        url = (info.get("hospUrl") or "").strip()
        if url and not url.lower().startswith(("http://", "https://")):
            url = "http://" + url
        hospital["homepage"] = url
        if not hospital["phone"]:
            hospital["phone"] = (info.get("telNo") or "").strip()
    except Exception as exc:  # noqa: BLE001 - 실패한 병원은 빈칸 유지
        print(f"  상세 조회 실패: {hospital['name']} ({exc})", flush=True)
    time.sleep(0.1)
    return hospital


def collect_details(session: requests.Session, hospitals: list[dict]) -> None:
    done = 0
    with ThreadPoolExecutor(max_workers=DETAIL_WORKERS) as pool:
        futures = [pool.submit(fetch_detail, session, h) for h in hospitals]
        for _ in as_completed(futures):
            done += 1
            if done % 100 == 0 or done == len(hospitals):
                print(f"  상세 {done}/{len(hospitals)}건 조회", flush=True)


def write_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "요양병원"

    columns = ["병원명", "지역", "전화번호", "이메일", "팩스", "홈페이지 주소"]
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    keys = ["name", "region", "phone", "email", "fax", "homepage"]
    for row_idx, h in enumerate(rows, start=2):
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=h[key])

    widths = {"A": 40, "B": 18, "C": 18, "D": 25, "E": 18, "F": 45}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    wb.save(OUTPUT_FILE)
    print(f"저장 완료: {OUTPUT_FILE} ({len(rows)}건)", flush=True)


if __name__ == "__main__":
    session = requests.Session()
    rows = collect_list(session)
    collect_details(session, rows)
    rows.sort(key=lambda h: (h["region"], h["name"]))
    with open(CACHE_FILE, "w", encoding="utf-8") as fp:
        json.dump(rows, fp, ensure_ascii=False)
    write_excel(rows)
